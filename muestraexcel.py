"""
================================================================================
VISUAL/EXCEL.PY — DASHBOARD QUANT EN EXCEL  (v7.1 — ANCHO 10 XL)
================================================================================

CAMBIOS v7.1:
  - Gráfico: anclado en col B, ancho 1000px (ancho 10)
  - Tabla:   columna R (índice 17, 0-based), fuente 13pt, fila 2x
  - Sin solape: chart 1000px, tabla desplazada a la derecha
================================================================================
"""

import os
import re
import math
import datetime
import logging
import bisect
from copy import deepcopy
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Protocol, Tuple
import csv

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

try:
    import xlsxwriter
    from xlsxwriter.utility import xl_col_to_name
    _HAS_XW = True
except ImportError:
    _HAS_XW = False

logger = logging.getLogger(__name__)

# ==============================================================================
# COLORES Y CONSTANTES
# ==============================================================================

COLORS = {
    "header_bg_metrics": "2D3436",
    "header_bg_params":  "636E72",
    "header_bg_id":      "2D3436",
    "text_white":        "FFFFFF",
    "text_dark":         "2D3436",
    "border_color":      "DFE6E9",
    "success_bg":        "F0FFF4",
    "danger_bg":         "FFF5F5",
    "accent_green":      "38A169",
    "accent_red":        "E53E3E",
    "row_alt":           "F8F9FA",
    "section_border":    "A0AEC0",
    "table_header_bg":   "EDF2F7",
}

FONT_TITLE = "Calibri"
FONT_BODY  = "Calibri"

METRICS_ORDER = [
    "TRADES_DIA", "LONG", "SHORT",
    "PROFIT_FACTOR", "WINRATE_PCT", "ROI_PCT", "MAX_DD_PCT", "EXPECTATIVA", "SHARPE",
    "PNL_NETO", "SALDO_ACTUAL", "COMISIONES_TOTAL",
]

# Alias y métricas extra para detectar correctamente columnas de métricas
# que llegan con variantes de nombre desde metrics.py/reporting.
METRIC_ALIASES_TO_CANONICAL = {
    "ROI": "ROI_PCT",
    "ROI%": "ROI_PCT",
    "RETURN": "ROI_PCT",
    "RETURN%": "ROI_PCT",
    "RETURN_PCT": "ROI_PCT",
    "WINRATE": "WINRATE_PCT",
    "WIN_RATE": "WINRATE_PCT",
    "WIN_RATE_PCT": "WINRATE_PCT",
    "WINRATE%": "WINRATE_PCT",
    "MAX_DRAWDOWN": "MAX_DD_PCT",
    "MAX_DRAWDOWN_PCT": "MAX_DD_PCT",
    "DD": "MAX_DD_PCT",
    "DD_PCT": "MAX_DD_PCT",
    "DRAWDOWN": "MAX_DD_PCT",
    "N_TRADES": "TOTAL_TRADES",
    "NUM_TRADES": "TOTAL_TRADES",
    "COUNT_TRADES": "TOTAL_TRADES",
    "TRADES_POR_DIA": "TRADES_DIA",
    "EXPECTANCY": "EXPECTATIVA",
    "SHARPE_RATIO": "SHARPE",
}

# Nombres para mostrar en las cabeceras del Excel (interno → display)
METRIC_DISPLAY_NAMES = {
    "TRADES_DIA":           "TRADES DIA",
    "LONG":                 "LONG",
    "SHORT":                "SHORT",
    # ── 6 métricas canónicas ─────────────────────────────────────────────
    "PROFIT_FACTOR":        "PROFIT FACTOR",
    "WINRATE_PCT":          "WIN RATE",
    "ROI_PCT":              "ROI",
    "MAX_DD_PCT":           "MAX DD",
    "EXPECTATIVA":          "EXPECTANCY",
    "SHARPE":               "SHARPE",
    # ── financiero ───────────────────────────────────────────────────────
    "PNL_NETO":             "BEN NETO",
    "SALDO_ACTUAL":         "SALDO ACTUAL",
    "COMISIONES_TOTAL":     "COMISIONES",
}

ALL_METRICS_ORDER = METRICS_ORDER

# Todas las métricas conocidas (para excluirlas de la sección PARÁMETROS)
ALL_KNOWN_METRICS = {
    "TOTAL_TRADES", "TRADES_DIA", "LONG", "SHORT",
    "PROFIT_FACTOR", "ROI_PCT", "WINRATE_PCT", "MAX_DD_PCT", "EXPECTATIVA", "SHARPE",
    "NET_PNL", "PNL_NETO", "SALDO_ACTUAL",
    "SALDO_MAX", "SALDO_MIN", "COMISIONES_TOTAL", "SALDO_SIN_COMISIONES",
    "ROI%", "WINRATE%", "MAX_DD%",
    "TRADES_POR_DIA", "DURATION_MEAN_MIN", "SALDO_MEAN",
}

ID_COLS = ["TRIAL", "ESTRATEGIA", "SCORE"]

EXCLUDED_PARAMS = {
    "NOMBRE_COMBO", "EXIT_TYPE", "CANTIDAD",
    "ACTIVO", "TIMEFRAME", "TF", "ASSET", "SYMBOL",
    "RESULTADO", "METRICS", "COMBO", "ESTATEGIA",
    "SALDO", "VOLUMEN", "APALANCAMIENTO",
}

METRIC_KEYWORDS_TO_DROP = [
    "PROFIT", "LOSS", "PNL", "NET", "GROSS", "SALDO", "BALANCE", "RETORNO", "RETURN",
    "ROI", "BENEFICIO", "RIESGO", "RISK", "REWARD", "COMISION", "FEES",
    "WIN", "GANADORA", "PERDEDORA", "ACIERTO", "RATE", "PCT", "PORC_", "PERCENT",
    "DRAWDOWN", "DD", "UNDERWATER",
    "RATIO", "FACTOR", "SHARPE", "EXPECTATIVA", "EXPECTANCY",
    "AVG", "MEAN", "MEDIAN", "STD", "VAR", "MAX", "MIN", "SUM", "TOTAL",
    "COUNT", "NUM_", "N_", "TRADES", "LONGS", "SHORTS", "CANTIDAD_OP",
    "METRIC", "RESULT", "BEST", "WORST", "DIA_OPERADO", "DURATION", "TIME"
]

PREFIXES_TO_CLEAN = [
    "ESTRATEGIA_PARAMS_", "STRATEGY_PARAMS_", "PARAM_", "PARAMS_",
    "INDICATOR_", "CONFIG_", "METRICS_"
]

# ==============================================================================
# EXCEL REPORTER
# ==============================================================================

class ReporterProtocol(Protocol):
    def needs_dataframe(self, score: float) -> bool: ...
    def on_trial_end(self, artifacts: Any) -> None: ...
    def on_strategy_end(self, strategy_name: str, study: Any) -> None: ...


@dataclass
class ExcelReporter:
    resumen_path: str = "resultados/excel/resumen.xlsx"
    trades_base_dir: str = "resultados/excel"
    max_archivos: int = 5
    use_fast_mode: bool = True
    # Datos de precio para la gráfica comparativa
    datos_dir: str = "datos"
    fecha_inicio: Optional[str] = None
    fecha_fin: Optional[str] = None
    formato_datos: str = "feather"

    comparativa_periodos_activar: bool = True
    comparativa_periodo: str = "1mes"

    # Régimen de mercado (para filtrar períodos inactivos)
    regimen_activo: bool = False
    regimen_tipo: str = "ALCISTA"
    regimen_buy_hold_activo: bool = False
    regimen_buy_hold_saldo: str = "ALL"  # "ALL" o importe fijo ("1500")

    _csv_resumen_path: Optional[str] = field(default=None, init=False, repr=False)
    _resumen_rows: List[Dict[str, Any]] = field(default_factory=list, init=False, repr=False)
    _trade_candidates: List[Dict[str, Any]] = field(default_factory=list, init=False, repr=False)
    _min_candidate_score: float = field(default=float("-inf"), init=False, repr=False)
    _activo: Optional[str] = field(default=None, init=False, repr=False)
    _final_excel_path: Optional[str] = field(default=None, init=False, repr=False)

    def needs_dataframe(self, score: float) -> bool:
        return False

    @staticmethod
    def _safe_activo_name(activo: str) -> str:
        return str(activo).strip().replace(" ", "_").upper() if activo else "DEFAULT"

    def _update_min_score(self):
        self._min_candidate_score = (
            min(c["score"] for c in self._trade_candidates)
            if self._trade_candidates else float("-inf")
        )

    def on_trial_end(self, artifacts) -> None:
        params_src = getattr(artifacts, "params_reporting", None) or artifacts.params
        activo = None
        if isinstance(params_src, dict):
            activo = params_src.get("__activo") or params_src.get("ACTIVO") or params_src.get("activo")

        self._activo = activo
        score  = artifacts.score if artifacts.score is not None else 0.0
        params = dict(params_src)
        params["NOMBRE_COMBO"] = artifacts.strategy_name

        self._resumen_rows.append({
            "trial_number":  artifacts.trial_number,
            "score":         score,
            "metrics":       deepcopy(artifacts.metrics) if artifacts.metrics else {},
            "params":        {k: v for k, v in params.items() if not str(k).startswith("__")},
            "strategy_name": artifacts.strategy_name,
        })

        try:
            base_dir = self.trades_base_dir
            os.makedirs(base_dir, exist_ok=True)
            if not self._csv_resumen_path:
                self._csv_resumen_path = os.path.join(base_dir, "RESUMEN.csv")
            self._write_resumen_csv(self._csv_resumen_path)
        except Exception:
            pass

        is_candidate = (
            len(self._trade_candidates) < self.max_archivos or
            score > self._min_candidate_score
        )
        if is_candidate and artifacts.trades is not None:
            self._trade_candidates.append({
                "score":        score,
                "trial_number": artifacts.trial_number,
                "trades":       artifacts.trades,
                "params":       params,
                "metrics":      artifacts.metrics,
            })
            if len(self._trade_candidates) > self.max_archivos:
                self._trade_candidates.sort(key=lambda x: x["score"], reverse=True)
                self._trade_candidates.pop()
            self._update_min_score()

    def on_strategy_end(self, strategy_name: str, study) -> None:
        if not self._resumen_rows:
            return
        import time as _t

        activo   = self._activo
        base_dir = self.trades_base_dir
        os.makedirs(base_dir, exist_ok=True)

        activo_safe = self._safe_activo_name(str(activo) if activo else "DEFAULT")
        csv_path    = os.path.join(base_dir, "RESUMEN.csv")
        self._write_resumen_csv(csv_path)

        # Cargar precio continuo UNA SOLA VEZ (cache para todos los trials)
        _t0 = _t.perf_counter()
        _cached_prices, _cached_timestamps = None, None
        _cached_prices_5k, _cached_ts_5k = None, None
        try:
            # 5000 pts para comparativa períodos (alta resolución)
            _cached_prices_5k, _cached_ts_5k = _load_price_with_timestamps(
                activo=self._activo or "BTC",
                datos_dir=self.datos_dir,
                fecha_inicio=self.fecha_inicio,
                fecha_fin=self.fecha_fin,
                n_points=5000,
            )
            # 500 pts para gráfica B&H (downsampled from the 5k)
            if _cached_prices_5k and len(_cached_prices_5k) > 500:
                n5k = len(_cached_prices_5k)
                idx500 = [int(round(i * (n5k - 1) / 499)) for i in range(500)]
                _cached_prices = [_cached_prices_5k[i] for i in idx500]
                _cached_timestamps = [_cached_ts_5k[i] for i in idx500]
            else:
                _cached_prices = _cached_prices_5k
                _cached_timestamps = _cached_ts_5k
        except Exception:
            pass
        _t1 = _t.perf_counter()

        logger.debug(f"[EXCEL DEBUG] Precio B&H cargado en {(_t1-_t0)*1000:.0f}ms ({len(_cached_prices) if _cached_prices else 0} pts)")

        # Pre-computar régimen diario UNA SOLA VEZ (para comparativa períodos)
        _cached_regime_daily = None
        if self.regimen_activo:
            try:
                import polars as _pl_regime
                from modelox.core.regime import compute_regime_mask_1d
                _act_up = str(self._activo or "BTC").strip().upper()
                _base = self.datos_dir or "datos"
                _price_path = None
                for _pat in [f"{_act_up}_ohlcv_1m.feather", f"{_act_up}_ohlcv_1m.parquet",
                             f"{_act_up.lower()}_ohlcv_1m.feather"]:
                    _p = os.path.join(_base, _pat)
                    if os.path.exists(_p):
                        _price_path = _p
                        break
                if _price_path:
                    _t_reg0 = _t.perf_counter()
                    _ext = os.path.splitext(_price_path)[1].lower()
                    if _ext in (".feather", ".fthr"):
                        _df_1m = _pl_regime.read_ipc(_price_path, memory_map=False)
                    else:
                        _df_1m = _pl_regime.read_parquet(_price_path)
                    _regime_df = compute_regime_mask_1d(_df_1m)
                    _cached_regime_daily = {}
                    for _row in _regime_df.iter_rows(named=True):
                        _d = str(_row["timestamp"])[:10]
                        _cached_regime_daily[_d] = _row["regime"]
                    del _df_1m, _regime_df
                    _t_reg1 = _t.perf_counter()
                    logger.debug(f"[EXCEL DEBUG] Régimen diario cacheado en {(_t_reg1-_t_reg0)*1000:.0f}ms ({len(_cached_regime_daily)} días)")
            except Exception as _e:
                logger.debug(f"[EXCEL DEBUG] Régimen no disponible: {_e}")

        self._trade_candidates.sort(key=lambda x: x["score"], reverse=True)
        for candidate in self._trade_candidates[:self.max_archivos]:
            try:
                _t2 = _t.perf_counter()
                self._write_trades_excel(base_dir, candidate,
                                         cached_prices=_cached_prices,
                                         cached_timestamps=_cached_timestamps,
                                         cached_prices_5k=_cached_prices_5k,
                                         cached_ts_5k=_cached_ts_5k,
                                         cached_regime_daily=_cached_regime_daily)
                _t3 = _t.perf_counter()
                logger.debug(f"[EXCEL DEBUG] Trial {candidate['trial_number']} Excel en {(_t3-_t2)*1000:.0f}ms")
            except Exception as e:
                logger.warning(f"Error guardando trades trial {candidate['trial_number']}: {e}")

        # Intentar leer fechas desde configuracion.py si no se pasaron
        _fecha_inicio = self.fecha_inicio
        _fecha_fin    = self.fecha_fin
        if _fecha_inicio is None:
            try:
                from general.configuracion import FECHA_INICIO, FECHA_FIN
                _fecha_inicio = FECHA_INICIO
                _fecha_fin    = FECHA_FIN
            except Exception:
                pass

        try:
            _t4 = _t.perf_counter()
            self._final_excel_path = convertir_resumen_csv_a_excel(
                csv_path=csv_path,
                strategy_name=strategy_name,
                activo=activo_safe,
                output_dir=base_dir,
                excel_path=self.resumen_path,
                datos_dir=self.datos_dir,
                fecha_inicio=_fecha_inicio,
                fecha_fin=_fecha_fin,
                formato_datos=self.formato_datos,
            )
            _t5 = _t.perf_counter()
            logger.debug(f"[EXCEL DEBUG] Resumen Excel en {(_t5-_t4)*1000:.0f}ms")
        except Exception as e:
            logger.warning(f"Error generando Dashboard Excel: {e}")

        self._resumen_rows        = []
        self._trade_candidates    = []
        self._min_candidate_score = float("-inf")

    def _write_resumen_csv(self, csv_path: str):
        if not self._resumen_rows:
            return
        all_keys = {"trial", "score", "strategy"}
        for row in self._resumen_rows:
            if row.get("metrics"):
                all_keys.update(row["metrics"].keys())
            if row.get("params"):
                all_keys.update(f"param_{k}" for k in row["params"].keys())

        columns = ["trial", "score", "strategy"]
        columns.extend(sorted(k for k in all_keys if k not in {"trial","score","strategy"} and not k.startswith("param_")))
        columns.extend(sorted(k for k in all_keys if k.startswith("param_")))

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            for row in self._resumen_rows:
                csv_row = {"trial": row["trial_number"], "score": row["score"], "strategy": row["strategy_name"]}
                if row.get("metrics"):
                    csv_row.update(row["metrics"])
                if row.get("params"):
                    csv_row.update({f"param_{k}": v for k, v in row["params"].items()})
                writer.writerow(csv_row)

    def _write_trades_excel(self, trades_dir: str, candidate: Dict[str, Any],
                            cached_prices=None, cached_timestamps=None,
                            cached_prices_5k=None, cached_ts_5k=None,
                            cached_regime_daily=None):
        import time as _t
        trades = candidate["trades"]
        if trades is None or (hasattr(trades, "empty") and trades.empty):
            return

        _t0 = _t.perf_counter()
        df_trades = trades.to_pandas() if hasattr(trades, "to_pandas") else trades

        # --- Limpiar timezone ---
        try:
            df_trades = df_trades.copy()
            if isinstance(df_trades.index, pd.DatetimeIndex) and df_trades.index.tz is not None:
                df_trades.index = df_trades.index.tz_localize(None)
            for col in list(df_trades.columns):
                try:
                    if isinstance(df_trades[col].dtype, pd.DatetimeTZDtype):
                        df_trades[col] = df_trades[col].dt.tz_localize(None)
                except Exception:
                    continue
            for col in list(df_trades.columns):
                if df_trades[col].dtype != object:
                    continue
                s = df_trades[col]
                try:
                    sample = next((v for v in s.head(50).tolist() if v is not None), None)
                    if sample and getattr(sample, "tzinfo", None):
                        df_trades[col] = s.apply(
                        )
                except Exception:
                    continue
        except Exception:
            pass
        _t1 = _t.perf_counter()

        # Desglose ROI por modo (Estrategia vs Buy&Hold) para tabla lateral
        regime_mode_breakdown = _calcular_aporte_roi_por_modo(
            df_trades,
            cached_price_vals=cached_prices_5k,
            cached_price_ts=cached_ts_5k,
        )

        exit_type = str(candidate["params"].get("__exit_type", candidate["params"].get("exit_type", "FIXED"))).upper()
        df_export = _preparar_df_trades(df_trades, exit_type)
        _t2 = _t.perf_counter()

        saldo = candidate['params'].get('__saldo_usado') or 0
        apal  = candidate['params'].get('__apalancamiento_max') or 0
        vol   = saldo * apal
        bh_saldo_str = str(self.regimen_buy_hold_saldo).strip().upper() if self.regimen_buy_hold_activo else None

        filename = f"TRIAL {candidate['trial_number']} - {int(candidate['score'])}.xlsx"
        filepath = os.path.join(trades_dir, filename)

        # ── Comparativa por períodos (opcional) ──────────────────────────────
        period_data = None
        if self.comparativa_periodos_activar:
            try:
                period_data = _calcular_comparativa_periodos(
                    df=df_export,
                    activo=self._activo or "BTC",
                    datos_dir=self.datos_dir,
                    fecha_inicio=self.fecha_inicio,
                    fecha_fin=self.fecha_fin,
                    periodo_str=self.comparativa_periodo,
                    regimen_activo=self.regimen_activo,
                    regimen_tipo=self.regimen_tipo,
                    cached_price_vals=cached_prices_5k,
                    cached_price_ts=cached_ts_5k,
                    cached_regime_daily=cached_regime_daily,
                )
            except Exception:
                pass

        try:
            _escribir_trades_xlsxwriter(
                filepath, df_export, saldo, vol, apal,
                price_series=cached_prices,
                price_timestamps=cached_timestamps,
                period_data=period_data,
                activo_name=str(self._activo or "Activo").upper(),
                regime_mode_breakdown=regime_mode_breakdown,
                bh_saldo_str=bh_saldo_str,
            )
        except Exception as e:
            logger.warning(f"Error xlsxwriter, fallback openpyxl: {e}")
            _escribir_trades_openpyxl_fallback(filepath, df_export, saldo, vol, apal)
        _t3 = _t.perf_counter()
        logger.debug(f"[EXCEL DEBUG]   prep={(_t1-_t0)*1000:.0f}ms  df={(_t2-_t1)*1000:.0f}ms  write={(_t3-_t2)*1000:.0f}ms  rows={len(df_export)}")


# ==============================================================================
# PREPARACIÓN DEL DATAFRAME
# ==============================================================================

def _preparar_df_trades(df_trades: pd.DataFrame, exit_type: str = "") -> pd.DataFrame:
    rename_map = {

        "entry_time": "ENTRY_TIME", "exit_time": "EXIT_TIME",
        "type": "POSICIÓN", "entry_price": "ENTRY_PRICE", "exit_price": "EXIT_PRICE",
        "qty": "CANTIDAD", "saldo_usado": "SALDO",
        "pnl_bruto": "PNL BRUTO", "comision": "COMISIONES",
        "pnl_neto": "PNL NETO", "pnl_pct": "ROI",
        "saldo_antes": "BALANCE_PRE", "saldo_despues": "BALANCE",
        "reason": "EXIT_REASON", "trail_act_price": "TRAIL_ACT_PRICE",
        "trail_act_time": "TRAIL_ACT_TIME"
    }
    df = df_trades.rename(columns=rename_map)

    cols_to_drop = [
        "ENTRY_IDX", "EXIT_IDX", "SIDE_INT", "entry_idx", "exit_idx", "side_int",
        "BALANCE_PRE", "TRAIL_ACT_IDX", "trail_act_idx",
        "ROI", "SALDO", "VOLUMEN", "APALANCAMIENTO"
    ]
    df.drop(columns=[c for c in cols_to_drop if c in df.columns], inplace=True)

    reason_map = {1: "SL", 2: "TP", 3: "TRAIL", 4: "TIME", 0: "END", 5: "CUSTOM"}
    if "EXIT_REASON" in df.columns:
        df["EXIT_REASON"] = df["EXIT_REASON"].map(reason_map).fillna("UNKNOWN")

    for c in ["POSICIÓN", "EXIT_REASON"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.upper()

    df.columns = [c.upper() for c in df.columns]

    # Columna TIPO: "B&H" si es buy&hold, "S" si es estrategia
    if "IS_BUY_HOLD" in df.columns:
        df["TIPO"] = df["IS_BUY_HOLD"].fillna(False).astype(bool).map({True: "B&H", False: "S"})
        df.drop(columns=["IS_BUY_HOLD"], inplace=True)
    else:
        df["TIPO"] = "S"

    # Volumen usado por trade (notional): CANTIDAD * ENTRY_PRICE
    if "CANTIDAD" in df.columns and "ENTRY_PRICE" in df.columns:
        try:
            _qty = pd.to_numeric(df["CANTIDAD"], errors="coerce")
            _epx = pd.to_numeric(df["ENTRY_PRICE"], errors="coerce")
            df["VOLUMEN USADO"] = (_qty * _epx).round(2)
        except Exception:
            pass

    order = [
        "TIPO",
        "ENTRY_TIME", "ENTRY_PRICE", "TRAIL_ACT_PRICE", "TRAIL_ACT_TIME",
        "EXIT_TIME", "EXIT_PRICE", "POSICIÓN", "EXIT_REASON", "CANTIDAD", "VOLUMEN USADO",
        "PNL BRUTO", "COMISIONES", "PNL NETO", "BALANCE"
    ]
    
    # Si la salida es FIXED (no TRAILING), no mostramos columnas de activación de Trail
    if "TRAIL" not in exit_type:
        order = [c for c in order if c not in ("TRAIL_ACT_PRICE", "TRAIL_ACT_TIME")]

    # Truncar precios a 2 decimales según petición del usuario

    for price_col in ["ENTRY_PRICE", "EXIT_PRICE", "TRAIL_ACT_PRICE"]:
        if price_col in df.columns:
            df[price_col] = df[price_col].apply(lambda x: int(x * 100) / 100.0 if pd.notnull(x) else x)

    return df[[c for c in order if c in df.columns]]


# ==============================================================================
# HELPER: major_unit "bonito" para el eje Y
# ==============================================================================

def _nice_major_unit(data_range: float, target_ticks: int = 6) -> float:
    """Calcula un major_unit legible para el eje Y."""
    if data_range <= 0:
        return 1.0
    raw = data_range / target_ticks
    magnitude = math.pow(10, math.floor(math.log10(raw)))
    normalized = raw / magnitude
    if normalized <= 1:
        nice = 1
    elif normalized <= 2:
        nice = 2
    elif normalized <= 5:
        nice = 5
    else:
        nice = 10
    return nice * magnitude


# ==============================================================================
# COMPARATIVA POR PERÍODOS — helpers
# ==============================================================================

def _parse_periodo_offset(periodo_str: str):
    """Parsea "1mes", "3 semanas", "2 años", "1 día" → pandas DateOffset."""
    import re
    from pandas.tseries.offsets import DateOffset

    s = str(periodo_str).strip().lower()
    m = re.match(
        r'(\d+)\s*'
        r'(dia|dias|day|days|semana|semanas|week|weeks|mes|meses|month|months|año|años|anos|year|years)',
        s
    )
    if not m:
        return DateOffset(months=1)

    n    = int(m.group(1))
    unit = m.group(2)

    if   unit in ('dia', 'dias', 'day', 'days'):
        return DateOffset(days=n)
    elif unit in ('semana', 'semanas', 'week', 'weeks'):
        return DateOffset(weeks=n)
    elif unit in ('mes', 'meses', 'month', 'months'):
        return DateOffset(months=n)
    elif unit in ('año', 'años', 'anos', 'year', 'years'):
        return DateOffset(years=n)
    return DateOffset(months=1)


def _periodo_label(t_start: "pd.Timestamp", offset) -> str:
    """Genera etiqueta legible para un período según su duración."""
    import pandas as pd
    from pandas.tseries.offsets import DateOffset

    # Detectar magnitud aproximada del offset en días
    try:
        delta_days = (t_start + offset - t_start).days
    except Exception:
        delta_days = 30

    MESES_ES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    if delta_days >= 300:          # ~1 año o más
        return str(t_start.year)
    elif delta_days >= 25:         # ~1 mes o más
        return f"{MESES_ES[t_start.month - 1]} {str(t_start.year)[2:]}"
    else:                          # semanas o días
        return t_start.strftime("%d/%m/%y")


def _calcular_comparativa_periodos(
    df: "pd.DataFrame",
    activo: str,
    datos_dir: str,
    fecha_inicio: Optional[str],
    fecha_fin: Optional[str],
    periodo_str: str,
    regimen_activo: bool = False,
    regimen_tipo: str = "ALCISTA",
    cached_price_vals: Optional[List[float]] = None,
    cached_price_ts: Optional[List] = None,
    cached_regime_daily: Optional[dict] = None,
) -> Optional[Tuple[List[str], List[float], List[float]]]:
    """
    Divide el rango de la estrategia en períodos y calcula la variación % de:
      - La estrategia (balance al cierre del período)
      - El activo subyacente (precio de cierre del período)

    Si regimen_activo=True, sólo incluye períodos donde el régimen dominante
    coincide con regimen_tipo.

    Retorna (labels, strat_pct_list, asset_pct_list) o None si no hay datos.
    """
    try:
        def _to_naive_ts(v):
            ts = pd.Timestamp(v)
            try:
                if ts.tzinfo is not None:
                    ts = ts.tz_convert(None)
            except Exception:
                try:
                    if getattr(ts, "tzinfo", None) is not None:
                        ts = ts.tz_localize(None)
                except Exception:
                    pass
            return ts

        if df is None or df.empty:
            return None

        # ── 1. Detectar columnas ─────────────────────────────────────────────
        cols_up   = {c.upper(): c for c in df.columns}
        exit_col  = cols_up.get("EXIT_TIME")  or cols_up.get("ENTRY_TIME")
        bal_col   = cols_up.get("BALANCE")
        pnl_col   = cols_up.get("PNL NETO")   or cols_up.get("PNL_NETO")

        if exit_col is None or bal_col is None or pnl_col is None:
            return None

        # ── 2. Construir curva de balance ordenada por tiempo ────────────────
        df_sorted = df[[exit_col, bal_col, pnl_col]].copy()
        df_sorted[exit_col] = pd.to_datetime(df_sorted[exit_col], errors='coerce')
        df_sorted = df_sorted.dropna(subset=[exit_col]).sort_values(exit_col)
        df_sorted = df_sorted[pd.to_numeric(df_sorted[bal_col], errors='coerce').notna()]
        df_sorted = df_sorted[pd.to_numeric(df_sorted[pnl_col], errors='coerce').notna()]

        if df_sorted.empty:
            return None

        # Saldo inicial = primer balance - primer pnl
        try:
            first_pnl = float(df_sorted[pnl_col].iloc[0])
            first_bal = float(df_sorted[bal_col].iloc[0])
            saldo_ini = first_bal - first_pnl
        except Exception:
            return None
        if saldo_ini <= 0:
            return None

        times_ser   = df_sorted[exit_col].map(_to_naive_ts)  # pd.Series de Timestamps naive
        balance_arr = df_sorted[bal_col].astype(float).values

        def _balance_at(t_pd: "pd.Timestamp") -> float:
            """Retorna el balance de la estrategia justo antes o en t_pd."""
            t_pd = pd.Timestamp(t_pd)
            idx = int((times_ser <= t_pd).sum()) - 1
            return float(balance_arr[idx]) if idx >= 0 else saldo_ini

        # ── 3. Rango de fechas del rango completo ────────────────────────────
        t_first = pd.Timestamp(times_ser.iloc[0])
        t_last  = pd.Timestamp(times_ser.iloc[-1])

        offset = _parse_periodo_offset(periodo_str)

        # ── 4. Cargar precios diarios completos para el activo ───────────────
        if cached_price_vals is not None and cached_price_ts is not None:
            price_vals, price_ts = cached_price_vals, cached_price_ts
        else:
            price_vals, price_ts = _load_price_with_timestamps(
                activo=activo or "BTC",
                datos_dir=datos_dir or "datos",
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                n_points=5000,
            )
        if not price_ts or len(price_ts) < 2:
            return None

        price_ts_pd   = [_to_naive_ts(t) for t in price_ts]
        price_vals_arr = list(price_vals)

        # Limitar el rango al solape real entre trades y precio del activo.
        t_first = max(t_first, price_ts_pd[0])
        t_last  = min(t_last, price_ts_pd[-1])
        if t_first >= t_last:
            return None

        def _price_at(t_pd: "pd.Timestamp") -> Optional[float]:
            """Precio más cercano en el tiempo (bisección)."""
            import bisect
            idx = bisect.bisect_right(price_ts_pd, t_pd) - 1
            if idx < 0:
                idx = 0
            if idx >= len(price_vals_arr):
                idx = len(price_vals_arr) - 1
            return float(price_vals_arr[idx])

        # ── 5. Cargar régimen si está activo ────────────────────────────────
        regime_daily = cached_regime_daily  # Usar cache si disponible
        if regimen_activo and regime_daily is None:
            try:
                import polars as _pl
                from modelox.core.regime import compute_regime_mask_1d
                _price_path = None
                import glob as _glob2
                _act_up = str(activo).strip().upper()
                _base = datos_dir or "datos"
                for _pat in [f"{_act_up}_ohlcv_1m.feather", f"{_act_up}_ohlcv_1m.parquet",
                             f"{_act_up.lower()}_ohlcv_1m.feather", f"{_act_up.lower()}_ohlcv_1m.parquet"]:
                    _p = os.path.join(_base, _pat)
                    if os.path.exists(_p):
                        _price_path = _p
                        break
                if _price_path:
                    _ext = os.path.splitext(_price_path)[1].lower()
                    if _ext in (".feather", ".fthr"):
                        _df_1m = _pl.read_ipc(_price_path, memory_map=False)
                    else:
                        _df_1m = _pl.read_parquet(_price_path)
                    _regime_df = compute_regime_mask_1d(_df_1m)
                    regime_daily = {}
                    for _row in _regime_df.iter_rows(named=True):
                        _d = str(_row["timestamp"])[:10]
                        regime_daily[_d] = _row["regime"]
                    del _df_1m
            except Exception as _e:
                logger.debug(f"_calcular_comparativa_periodos régimen: {_e}")
                regime_daily = None

        def _period_regime_ok(t_start, t_end):
            """Retorna True si el régimen dominante en [t_start, t_end) coincide con regimen_tipo."""
            if regime_daily is None:
                return True
            _target = regimen_tipo.upper()
            _match = 0
            _total = 0
            _d = t_start.normalize()
            while _d < t_end:
                _key = _d.strftime("%Y-%m-%d")
                if _key in regime_daily:
                    _total += 1
                    if regime_daily[_key] == _target:
                        _match += 1
                _d += pd.Timedelta(days=1)
            if _total == 0:
                return True
            return _match / _total > 0.5  # >50% del período en el régimen correcto

        # ── 6. Iterar períodos ───────────────────────────────────────────────
        labels:     List[str]   = []
        strat_pcts: List[float] = []
        asset_pcts: List[float] = []

        t_cur = t_first.normalize()   # alinear al inicio del día
        while t_cur < t_last:
            t_next = t_cur + offset
            if t_next > t_last:
                t_next = t_last

            # Filtrar por régimen: saltar períodos donde la estrategia no operaba
            if not _period_regime_ok(t_cur, t_next):
                t_cur = t_next
                continue

            bal_start = _balance_at(t_cur)
            bal_end   = _balance_at(t_next)

            p_start = _price_at(t_cur)
            p_end   = _price_at(t_next)

            if bal_start > 0 and p_start and p_start > 0:
                strat_pct = round((bal_end / bal_start - 1.0) * 100.0, 2)
                asset_pct = round((p_end   / p_start - 1.0) * 100.0, 2)

                labels.append(_periodo_label(t_cur, offset))
                strat_pcts.append(strat_pct)
                asset_pcts.append(asset_pct)

            t_cur = t_next

        if len(labels) < 2:
            return None

        return labels, strat_pcts, asset_pcts

    except Exception as _e:
        logger.debug(f"_calcular_comparativa_periodos: {_e}")
        return None


def _calcular_aporte_roi_por_modo(
    df_trades: pd.DataFrame,
    cached_price_vals: Optional[List[float]] = None,
    cached_price_ts: Optional[List] = None,
) -> Optional[Dict[str, float]]:
    """
    Calcula aporte de ROI por modo (Estrategia vs Buy&Hold por régimen) y
    rendimiento del activo en los períodos operados por cada modo.
    """
    try:
        if df_trades is None or df_trades.empty:
            return None

        req = {"pnl_neto", "entry_time", "exit_time"}
        if not req.issubset(set(df_trades.columns)):
            return None

        if "is_buy_hold" not in df_trades.columns:
            return None

        dfx = df_trades.copy()
        dfx["pnl_neto"] = pd.to_numeric(dfx["pnl_neto"], errors="coerce").fillna(0.0)
        dfx["is_buy_hold"] = dfx["is_buy_hold"].fillna(False).astype(bool)

        if "saldo_antes" in dfx.columns:
            saldo_ini = float(pd.to_numeric(dfx["saldo_antes"], errors="coerce").dropna().iloc[0])
        elif "saldo_despues" in dfx.columns:
            first_bal = float(pd.to_numeric(dfx["saldo_despues"], errors="coerce").dropna().iloc[0])
            first_pnl = float(dfx["pnl_neto"].iloc[0])
            saldo_ini = first_bal - first_pnl
        else:
            return None

        if saldo_ini <= 0:
            return None

        pnl_total = float(dfx["pnl_neto"].sum())
        pnl_bh = float(dfx.loc[dfx["is_buy_hold"], "pnl_neto"].sum())
        pnl_strat = float(dfx.loc[~dfx["is_buy_hold"], "pnl_neto"].sum())

        roi_total = (pnl_total / saldo_ini) * 100.0
        roi_bh = (pnl_bh / saldo_ini) * 100.0
        roi_strat = (pnl_strat / saldo_ini) * 100.0

        if abs(roi_total) > 1e-12:
            share_bh = (roi_bh / roi_total) * 100.0
            share_strat = (roi_strat / roi_total) * 100.0
        else:
            share_bh = 0.0
            share_strat = 0.0

        # Rendimiento del activo en TODO el periodo del backtest
        # (inicio→fin del rango de precios cacheado), para evitar sesgo
        # de calcular sólo cuando hubo trades.
        asset_ret_total = 0.0

        # Rendimiento del activo en los períodos operados por cada modo
        # (se mantiene como métrica auxiliar).
        asset_ret_bh = 0.0
        asset_ret_strat = 0.0
        if cached_price_vals and cached_price_ts and len(cached_price_vals) >= 2:
            pts = [pd.Timestamp(t).tz_localize(None) if getattr(pd.Timestamp(t), "tzinfo", None) else pd.Timestamp(t)
                   for t in cached_price_ts]
            pvs = [float(x) for x in cached_price_vals]

            def _price_at(ts_val: pd.Timestamp) -> Optional[float]:
                i = bisect.bisect_right(pts, ts_val) - 1
                if i < 0:
                    i = 0
                if i >= len(pvs):
                    i = len(pvs) - 1
                p = pvs[i]
                return p if p > 0 else None

            # ROI del activo en todo el backtest (primer precio vs último)
            p_start_all = pvs[0] if pvs and pvs[0] > 0 else None
            p_end_all = pvs[-1] if pvs and pvs[-1] > 0 else None
            if p_start_all and p_end_all:
                asset_ret_total = (p_end_all / p_start_all - 1.0) * 100.0

            def _asset_perf(mask: pd.Series) -> float:
                growth = 1.0
                sub = dfx.loc[mask, ["entry_time", "exit_time"]].copy()
                if sub.empty:
                    return 0.0
                sub["entry_time"] = pd.to_datetime(sub["entry_time"], errors="coerce")
                sub["exit_time"] = pd.to_datetime(sub["exit_time"], errors="coerce")
                sub = sub.dropna(subset=["entry_time", "exit_time"])
                for _, r in sub.iterrows():
                    et = r["entry_time"]
                    xt = r["exit_time"]
                    if getattr(et, "tzinfo", None):
                        et = et.tz_localize(None)
                    if getattr(xt, "tzinfo", None):
                        xt = xt.tz_localize(None)
                    p0 = _price_at(pd.Timestamp(et))
                    p1 = _price_at(pd.Timestamp(xt))
                    if p0 and p1 and p0 > 0:
                        growth *= (p1 / p0)
                return (growth - 1.0) * 100.0

            asset_ret_bh = _asset_perf(dfx["is_buy_hold"])
            asset_ret_strat = _asset_perf(~dfx["is_buy_hold"])

        n_strat = int((~dfx["is_buy_hold"]).sum())
        n_bh    = int(dfx["is_buy_hold"].sum())

        return {
            "has_buy_hold":    bool(dfx["is_buy_hold"].any()),
            "roi_total":       float(roi_total),
            "roi_strat":       float(roi_strat),
            "roi_bh":          float(roi_bh),
            "pnl_strat":       float(pnl_strat),
            "pnl_bh":          float(pnl_bh),
            "n_strat":         n_strat,
            "n_bh":            n_bh,
            "share_roi_strat": float(share_strat),
            "share_roi_bh":    float(share_bh),
            "asset_ret_total": float(asset_ret_total),
            "asset_ret_strat": float(asset_ret_strat),
            "asset_ret_bh":    float(asset_ret_bh),
        }
    except Exception:
        return None


# ==============================================================================
# ESCRITURA ULTRA RÁPIDA CON XLSXWRITER  (v7 — posición y escalado corregidos)
# ==============================================================================

def _escribir_trades_xlsxwriter(
    filepath: str,
    df: pd.DataFrame,
    val_saldo: float = 0,
    val_volumen: float = 0,
    val_apal: float = 0,
    price_series: Optional[List[float]] = None,
    price_timestamps: Optional[List] = None,
    period_data: Optional[Tuple[List[str], List[float], List[float]]] = None,
    activo_name: str = "Activo",
    regime_mode_breakdown: Optional[Dict[str, float]] = None,
    bh_saldo_str: Optional[str] = None,  # "ALL" o importe fijo; None = sin B&H
):
    """
    Layout final:
      · Datos:  A1 → L(max_row+1)
      · Gráfico: anclado en B(max_row+3), ancho 500 px × alto 430 px
                 → ocupa aproximadamente B → I (8 cols × ~65 px = 520 px)
      · Tabla:  anclada en J(max_row+3), inmediatamente junto al gráfico
                fuente 13 pt, filas 28 px (tamaño 2x)
    """
    n_rows = len(df)
    cols   = list(df.columns)

    wb = xlsxwriter.Workbook(filepath, {
        'nan_inf_to_errors': True,
        'strings_to_numbers': False,
        # Necesario para que xlsxwriter serialice datetime correctamente
        'default_date_format': 'dd/mm/yy hh:mm',
    })
    ws = wb.add_worksheet('Trades')
    ws.hide_gridlines(2)
    ws.freeze_panes(1, 0)
    ws.set_row(0, 28)

    # ── BASE de formato ──────────────────────────────────────────────────────
    _BASE = dict(
        font_name='Calibri', font_size=10, font_color='#2D3436',
        align='center', valign='vcenter',
        bottom=1, bottom_color='#E2E8F0',
        top=0, left=0, right=0,
    )

    def _fmt(bg='#FFFFFF', num_format=None, **kw):
        p = {**_BASE, 'bg_color': bg}
        if num_format:
            p['num_format'] = num_format
        p.update(kw)
        return wb.add_format(p)

    hdr_fmt = wb.add_format({
        'bold': True, 'font_name': 'Calibri', 'font_size': 10,
        'font_color': '#718096', 'bg_color': '#F7FAFC',
        'align': 'center', 'valign': 'vcenter',
        'bottom': 2, 'bottom_color': '#A0AEC0',
        'top': 0, 'left': 0, 'right': 0,
        'text_wrap': True,
    })

    # Pares de formato (fila normal, fila alternada)
    FMT = {
        'gen':   (_fmt('#FFFFFF'),                          _fmt('#F7FAFC')),
        'dt':    (_fmt('#FFFFFF', 'dd/mm/yy hh:mm'),        _fmt('#F7FAFC', 'dd/mm/yy hh:mm')),
        'price': (_fmt('#FFFFFF', '#,##0.00'),               _fmt('#F7FAFC', '#,##0.00')),
        'money': (_fmt('#FFFFFF', '#,##0.00'),               _fmt('#F7FAFC', '#,##0.00')),
        'pct':   (_fmt('#FFFFFF', '0.00%'),                  _fmt('#F7FAFC', '0.00%')),
        'apal':  (_fmt('#FFFFFF', '0.00"x"'),                _fmt('#F7FAFC', '0.00"x"')),
    }

    def _fmt_key(hdr: str) -> str:
        h = hdr.upper()
        if 'TIME' in h or 'DATE' in h:                          return 'dt'
        if 'PRICE' in h:                                         return 'price'
        if 'PNL' in h or 'BALANCE' in h or 'COMISIONES' in h or 'VOLUMEN' in h:  return 'money'
        if h == 'ROI':                                           return 'pct'
        if 'APALANCAMIENTO' in h:                                return 'apal'
        return 'gen'

    col_keys = [_fmt_key(c) for c in cols]

    # Índices de columnas especiales
    pnl_neto_idx    = next((i for i, c in enumerate(cols) if 'PNL' in c and 'NETO' in c), None)
    balance_idx     = next((i for i, c in enumerate(cols) if c == 'BALANCE'), None)
    entry_time_idx  = next((i for i, c in enumerate(cols) if c == 'ENTRY_TIME'), None)
    comisiones_idx  = next((i for i, c in enumerate(cols) if 'COMISIONES' in c), None)
    entry_price_idx = next((i for i, c in enumerate(cols) if c == 'ENTRY_PRICE'), None)
    exit_price_idx  = next((i for i, c in enumerate(cols) if c == 'EXIT_PRICE'), None)
    exit_time_idx   = next((i for i, c in enumerate(cols) if c == 'EXIT_TIME'), None)
    qty_idx         = next((i for i, c in enumerate(cols) if c == 'CANTIDAD'), None)
    posicion_idx    = next((i for i, c in enumerate(cols) if c == 'POSICIÓN'), None)
    tipo_idx        = next((i for i, c in enumerate(cols) if c == 'TIPO'), None)

    # ── Hoja auxiliar oculta para datos de gráficos ─────────────────────
    # Columnas: A=ENTRY_TIME, B=BAL_BRUTO, C=STRAT_ROI%, D=BH_ROI%
    #           E=CUM_COMISIONES, F=CUM_PNL_NETO
    #           G=BH_CONT_TIME, H=BH_CONT_ROI% (B&H continuo, independiente de trades)
    #           I=DAILY_TIME, J=DAILY_BAL, K=DAILY_ROI, L=DAILY_PNL, M=DAILY_BAL_BRUTO
    #           N=DAILY_BAL_STRAT (solo periodos estrategia), O=DAILY_BAL_BH (solo periodos B&H)
    #           P=DAILY_ROI_STRAT (solo periodos estrategia), Q=DAILY_ROI_BH (solo periodos B&H)
    has_bruto_data = False
    has_roi_data   = False
    has_continuous_bh = False
    has_daily_bal  = False
    n_bh_points = 0
    n_daily_points = 0
    if n_rows > 0:
        ws_aux = wb.add_worksheet('_ChartData')
        ws_aux.hide()
        ws_aux.write_string(0, 0, 'ENTRY_TIME')
        ws_aux.write_string(0, 1, 'BAL_BRUTO')
        ws_aux.write_string(0, 2, 'STRAT_IDX')
        ws_aux.write_string(0, 3, 'BH_IDX')
        ws_aux.write_string(0, 4, 'CUM_COMISIONES')
        ws_aux.write_string(0, 5, 'CUM_PNL_NETO')
        ws_aux.write_string(0, 6, 'BH_CONT_TIME')
        ws_aux.write_string(0, 7, 'BH_CONT_ROI')
        ws_aux.write_string(0, 8, 'DAILY_TIME')
        ws_aux.write_string(0, 9, 'DAILY_BAL')
        ws_aux.write_string(0, 10, 'DAILY_ROI')
        ws_aux.write_string(0, 11, 'DAILY_PNL')
        ws_aux.write_string(0, 12, 'DAILY_BAL_BRUTO')
        ws_aux.write_string(0, 13, 'DAILY_BAL_STRAT')
        ws_aux.write_string(0, 14, 'DAILY_BAL_BH')
        ws_aux.write_string(0, 15, 'DAILY_ROI_STRAT')
        ws_aux.write_string(0, 16, 'DAILY_ROI_BH')

        dt_fmt = wb.add_format({'num_format': 'dd/mm/yy hh:mm'})

        # Primer precio del activo (para Buy & Hold fallback)
        first_entry_price = None
        _saldo_ini = 0.0  # Se calcula en la primera iteración
        if entry_price_idx is not None:
            try:
                first_entry_price = float(df.iloc[0, entry_price_idx])
            except (ValueError, TypeError):
                first_entry_price = None

        comisiones_acum = 0.0
        pnl_neto_acum   = 0.0
        has_fee_data    = False
        for r_idx in range(n_rows):
            # Col A: Timestamp
            if entry_time_idx is not None:
                try:
                    et_val = df.iloc[r_idx, entry_time_idx]
                    if isinstance(et_val, (pd.Timestamp, datetime.datetime)):
                        dt = et_val.to_pydatetime() if hasattr(et_val, 'to_pydatetime') else et_val
                        dt = dt.replace(tzinfo=None)
                        ws_aux.write_datetime(r_idx + 1, 0, dt, dt_fmt)
                    else:
                        ws_aux.write(r_idx + 1, 0, str(et_val))
                except Exception:
                    ws_aux.write(r_idx + 1, 0, r_idx)

            # Col B: Balance Bruto
            if balance_idx is not None and comisiones_idx is not None:
                try:
                    comision_val = float(df.iloc[r_idx, comisiones_idx])
                except (ValueError, TypeError):
                    comision_val = 0.0
                comisiones_acum += comision_val
                try:
                    balance_val = float(df.iloc[r_idx, balance_idx])
                except (ValueError, TypeError):
                    balance_val = 0.0
                bruto_val = balance_val + comisiones_acum
                ws_aux.write_number(r_idx + 1, 1, bruto_val)
                has_bruto_data = True

            # Col C: Strategy ROI % (0% = punto de partida)
            if balance_idx is not None and pnl_neto_idx is not None:
                try:
                    bal = float(df.iloc[r_idx, balance_idx])
                except (ValueError, TypeError):
                    bal = 0.0
                if r_idx == 0:
                    try:
                        first_pnl = float(df.iloc[0, pnl_neto_idx])
                        first_bal = float(df.iloc[0, balance_idx])
                        _saldo_ini = first_bal - first_pnl
                    except (ValueError, TypeError):
                        _saldo_ini = first_bal if first_bal > 0 else 1.0
                if _saldo_ini > 0:
                    strat_roi = (bal / _saldo_ini - 1.0) * 100.0
                    ws_aux.write_number(r_idx + 1, 2, strat_roi)
                    has_roi_data = True

            # Col D: Buy & Hold ROI % (fallback: basado en trades)
            if not price_series:  # Solo si NO hay price_series continuo
                if entry_price_idx is not None and first_entry_price and first_entry_price > 0:
                    if r_idx == n_rows - 1 and exit_price_idx is not None:
                        try:
                            price_now = float(df.iloc[r_idx, exit_price_idx])
                        except (ValueError, TypeError):
                            price_now = first_entry_price
                    else:
                        try:
                            price_now = float(df.iloc[r_idx, entry_price_idx])
                        except (ValueError, TypeError):
                            price_now = first_entry_price
                    bh_roi = (price_now / first_entry_price - 1.0) * 100.0
                    ws_aux.write_number(r_idx + 1, 3, bh_roi)

            # Col E: Comisiones acumuladas
            if comisiones_idx is not None:
                try:
                    com_val = float(df.iloc[r_idx, comisiones_idx])
                except (ValueError, TypeError):
                    com_val = 0.0
                ws_aux.write_number(r_idx + 1, 4, comisiones_acum)
                has_fee_data = True

            # Col F: PNL Neto acumulado
            if pnl_neto_idx is not None:
                try:
                    pnl_val = float(df.iloc[r_idx, pnl_neto_idx])
                except (ValueError, TypeError):
                    pnl_val = 0.0
                pnl_neto_acum += pnl_val
                ws_aux.write_number(r_idx + 1, 5, pnl_neto_acum)

        # ── Col G/H: Buy & Hold CONTINUO (precio real, independiente de trades) ──
        if price_series and len(price_series) >= 2:
            n_bh_points = len(price_series)
            first_price = price_series[0]
            has_continuous_bh = first_price > 0

            if has_continuous_bh:
                # Escribir timestamps continuos (col G) y ROI% (col H)
                if price_timestamps and len(price_timestamps) == n_bh_points:
                    for bi in range(n_bh_points):
                        try:
                            ts_val = price_timestamps[bi]
                            if isinstance(ts_val, (pd.Timestamp, datetime.datetime)):
                                dt_bh = ts_val.to_pydatetime() if hasattr(ts_val, 'to_pydatetime') else ts_val
                                dt_bh = dt_bh.replace(tzinfo=None)
                                ws_aux.write_datetime(bi + 1, 6, dt_bh, dt_fmt)
                            else:
                                ws_aux.write(bi + 1, 6, str(ts_val))
                        except Exception:
                            ws_aux.write_number(bi + 1, 6, bi)
                        bh_roi_cont = (price_series[bi] / first_price - 1.0) * 100.0
                        ws_aux.write_number(bi + 1, 7, bh_roi_cont)
                else:
                    # Sin timestamps — usar índices
                    for bi in range(n_bh_points):
                        ws_aux.write_number(bi + 1, 6, bi)
                        bh_roi_cont = (price_series[bi] / first_price - 1.0) * 100.0
                        ws_aux.write_number(bi + 1, 7, bh_roi_cont)

        # ── Col I/J: Balance diario mark-to-market ──────────────────────────
        # Interpola el balance en cada punto de price_series teniendo en cuenta
        # el PnL no realizado de cualquier trade abierto en ese momento.
        # Esto elimina los escalones cuando hay trades largos (ej. B&H).
        if (price_series and price_timestamps
                and len(price_series) >= 2
                and entry_time_idx is not None
                and exit_time_idx is not None
                and qty_idx is not None
                and balance_idx is not None
                and pnl_neto_idx is not None
                and entry_price_idx is not None):

            def _ts_naive(v):
                ts = pd.Timestamp(v)
                return ts.tz_localize(None) if ts.tzinfo is not None else ts

            # Construir lista de trades ordenados por entry_time
            # Tupla: (entry_time, exit_time, entry_price, qty, pnl_neto, balance, posicion, comision, is_bh)
            _trades_mtm = []
            for _r in range(n_rows):
                try:
                    _et  = _ts_naive(df.iloc[_r, entry_time_idx])
                    _xt  = _ts_naive(df.iloc[_r, exit_time_idx])
                    _ep  = float(df.iloc[_r, entry_price_idx])
                    _qty = float(df.iloc[_r, qty_idx])
                    _pnl = float(df.iloc[_r, pnl_neto_idx])
                    _bal = float(df.iloc[_r, balance_idx])
                    _pos = str(df.iloc[_r, posicion_idx]).upper() if posicion_idx is not None else 'LONG'
                    _com = float(df.iloc[_r, comisiones_idx]) if comisiones_idx is not None else 0.0
                    _ibh = (str(df.iloc[_r, tipo_idx]).strip() == 'B&H') if tipo_idx is not None else False
                    _trades_mtm.append((_et, _xt, _ep, _qty, _pnl, _bal, _pos, _com, _ibh))
                except Exception:
                    continue
            _trades_mtm.sort(key=lambda x: x[0])

            # Saldo inicial (antes del primer trade)
            _saldo_ini_mtm = (_trades_mtm[0][5] - _trades_mtm[0][4]
                              if _trades_mtm else (_saldo_ini if _saldo_ini > 0 else 300.0))

            _last_closed_bal = _saldo_ini_mtm
            _cum_com = 0.0   # comisiones acumuladas de trades cerrados
            _ti = 0
            _nt = len(_trades_mtm)
            _daily_times:   list = []
            _daily_bals:    list = []
            _daily_rois:    list = []
            _daily_pnls:    list = []
            _daily_brutos:  list = []
            _daily_is_bh:   list = []   # True = B&H, False = Estrategia
            _prev_is_bh: object = None  # para detectar transiciones

            for _price_now, _ts_raw in zip(price_series, price_timestamps):
                try:
                    _ts = _ts_naive(pd.Timestamp(_ts_raw))
                except Exception:
                    continue

                # Avanzar trades ya cerrados en este momento
                while _ti < _nt and _trades_mtm[_ti][1] <= _ts:
                    _last_closed_bal = _trades_mtm[_ti][5]
                    _cum_com += _trades_mtm[_ti][7]
                    _ti += 1

                # ¿Hay un trade abierto ahora?
                if _ti < _nt and _trades_mtm[_ti][0] <= _ts:
                    _et, _xt, _ep, _qty, _pnl, _bal, _pos, _com, _ibh = _trades_mtm[_ti]
                    _bal_before = _bal - _pnl
                    _mtm_pnl = _qty * (_ep - _price_now) if 'SHORT' in _pos else _qty * (_price_now - _ep)
                    _cur_bal  = _bal_before + _mtm_pnl
                    _cur_is_bh = _ibh
                else:
                    _cur_bal   = _last_closed_bal
                    _cur_is_bh = False   # sin trade abierto → periodo de estrategia

                _cur_bruto = _cur_bal + _cum_com

                _daily_times.append(_ts_raw)
                _daily_bals.append(_cur_bal)
                _daily_rois.append((_cur_bal / _saldo_ini_mtm - 1.0) * 100.0 if _saldo_ini_mtm > 0 else 0.0)
                _daily_pnls.append(_cur_bal - _saldo_ini_mtm)
                _daily_brutos.append(_cur_bruto)
                _daily_is_bh.append(_cur_is_bh)
                _prev_is_bh = _cur_is_bh

            # Escribir en _ChartData cols I-O
            n_daily_points = len(_daily_times)
            if n_daily_points >= 2:
                _prev_bh_write = None
                for _di, (_dt_raw, _db, _dr, _dp, _dg, _dibh) in enumerate(
                    zip(_daily_times, _daily_bals, _daily_rois, _daily_pnls, _daily_brutos, _daily_is_bh)
                ):
                    try:
                        _dt_wr = pd.Timestamp(_dt_raw).to_pydatetime()
                        _dt_wr = _dt_wr.replace(tzinfo=None)
                        ws_aux.write_datetime(_di + 1, 8, _dt_wr, dt_fmt)
                    except Exception:
                        ws_aux.write_number(_di + 1, 8, _di)
                    ws_aux.write_number(_di + 1, 9,  float(_db))   # DAILY_BAL
                    ws_aux.write_number(_di + 1, 10, float(_dr))   # DAILY_ROI
                    ws_aux.write_number(_di + 1, 11, float(_dp))   # DAILY_PNL
                    ws_aux.write_number(_di + 1, 12, float(_dg))   # DAILY_BAL_BRUTO
                    # Cols N/O y P/Q: split por tipo — overlap de 1 punto en transiciones
                    _at_trans = (_prev_bh_write is not None) and (_dibh != _prev_bh_write)
                    if _dibh:
                        ws_aux.write_number(_di + 1, 14, float(_db))   # DAILY_BAL_BH
                        ws_aux.write_number(_di + 1, 16, float(_dr))   # DAILY_ROI_BH
                        if _at_trans:
                            ws_aux.write_number(_di + 1, 13, float(_db))  # overlap → conecta
                            ws_aux.write_number(_di + 1, 15, float(_dr))  # overlap → conecta
                    else:
                        ws_aux.write_number(_di + 1, 13, float(_db))   # DAILY_BAL_STRAT
                        ws_aux.write_number(_di + 1, 15, float(_dr))   # DAILY_ROI_STRAT
                        if _at_trans:
                            ws_aux.write_number(_di + 1, 14, float(_db))  # overlap → conecta
                            ws_aux.write_number(_di + 1, 16, float(_dr))  # overlap → conecta
                    _prev_bh_write = _dibh
                has_daily_bal = True

    # ── Anchos de columna: muestrea 20 filas, O(cols) ───────────────────────
    for i, col_name in enumerate(cols):
        max_len = len(str(col_name))
        if n_rows > 0:
            sample_len = df.iloc[:20, i].astype(str).str.len().max()
            if pd.notna(sample_len):
                max_len = max(max_len, int(sample_len))
        ws.set_column(i, i, min((max_len + 2) * 1.15, 28))

    # ── HEADERS ─────────────────────────────────────────────────────────────
    for i, col_name in enumerate(cols):
        ws.write(0, i, col_name, hdr_fmt)

    # ── DATOS: O(n) en C nativo, ~20-50x más rápido que openpyxl ───────────
    for r_idx, row_data in enumerate(df.itertuples(index=False), start=1):
        alt = (r_idx % 2 == 0)
        ws.set_row(r_idx, 18)

        for c_idx, val in enumerate(row_data):
            fmt = FMT[col_keys[c_idx]][1 if alt else 0]

            if val is None or (isinstance(val, float) and pd.isna(val)):
                ws.write_blank(r_idx, c_idx, None, fmt)
            elif isinstance(val, (pd.Timestamp, datetime.datetime)):
                try:
                    dt = val.to_pydatetime() if hasattr(val, 'to_pydatetime') else val
                    dt = dt.replace(tzinfo=None)
                    ws.write_datetime(r_idx, c_idx, dt, fmt)
                except Exception:
                    ws.write_string(r_idx, c_idx, str(val), fmt)
            elif isinstance(val, bool):
                ws.write_boolean(r_idx, c_idx, val, fmt)
            elif isinstance(val, (int, float)):
                ws.write_number(r_idx, c_idx, float(val), fmt)
            else:
                ws.write_string(r_idx, c_idx, str(val), fmt)

    # ── FORMATO CONDICIONAL PNL NETO ────────────────────────────────────────
    if pnl_neto_idx is not None and n_rows > 0:
        fmt_green = wb.add_format({**_BASE, 'font_color': '#38A169', 'bold': True})
        fmt_red   = wb.add_format({**_BASE, 'font_color': '#E53E3E', 'bold': True})
        ws.conditional_format(1, pnl_neto_idx, n_rows, pnl_neto_idx,
                               {'type': 'cell', 'criteria': '>', 'value': 0, 'format': fmt_green})
        ws.conditional_format(1, pnl_neto_idx, n_rows, pnl_neto_idx,
                               {'type': 'cell', 'criteria': '<', 'value': 0, 'format': fmt_red})

    # ── FORMATO CONDICIONAL TIPO (S / B&H) ──────────────────────────────────
    tipo_idx = next((i for i, c in enumerate(cols) if c == 'TIPO'), None)
    if tipo_idx is not None and n_rows > 0:
        fmt_bh = wb.add_format({**_BASE, 'bg_color': '#EBF8FF', 'font_color': '#2B6CB0', 'bold': True})
        ws.conditional_format(1, tipo_idx, n_rows, tipo_idx,
                               {'type': 'text', 'criteria': 'containing',
                                'value': 'B&H', 'format': fmt_bh})

    # ── FILA DE SEPARACIÓN entre datos y gráfico/tabla ──────────────────────
    BLOCK_ROW = n_rows + 2   # fila 0-indexed donde empieza bloque inferior

    CHARTS_HEIGHT_ROWS = 0  # filas que ocupan los gráficos (para calcular TABLE_ROW)

    # ── Tema visual unificado (look más profesional) ────────────────────────
    _CHART_TITLE_FONT = {'size': 12, 'bold': True, 'color': '#1F2937', 'name': 'Calibri'}
    _CHART_AXIS_Y = {
        'num_font':        {'size': 9, 'color': '#6B7280', 'name': 'Calibri'},
        'major_gridlines': {'visible': True, 'line': {'color': '#E5E7EB', 'width': 0.8}},
        'minor_gridlines': {'visible': False},
        'line':            {'none': True},
    }
    _CHART_AXIS_X = {
        'num_font':        {'size': 8, 'color': '#9CA3AF', 'name': 'Calibri'},
        'major_gridlines': {'visible': False},
        'major_tick_mark': 'none',
        'minor_tick_mark': 'none',
        'line':            {'color': '#E5E7EB', 'width': 0.75},
        'label_position':  'low',
    }
    _CHART_LEGEND = {'position': 'bottom', 'font': {'size': 9, 'color': '#6B7280', 'name': 'Calibri'}}
    _CHART_PLOTAREA = {'border': {'none': True}, 'fill': {'color': '#FFFFFF'}}
    _CHART_CHARTAREA = {'border': {'none': True}, 'fill': {'color': '#FCFCFD'}}
    _CHART_SIZE = {'width': 1040, 'height': 420}

    # ── GRÁFICO 1: EVOLUCIÓN DEL BALANCE (solo Balance Neto) ────────────────
    if balance_idx is not None and n_rows > 0:
        bal_series = df.iloc[:, balance_idx].dropna()
        y_min_raw  = float(bal_series.min())
        y_max_raw  = float(bal_series.max())
        data_range = y_max_raw - y_min_raw if y_max_raw != y_min_raw else max(abs(y_max_raw) * 0.1, 1.0)

        margin     = data_range * 0.03
        y_min_axis = y_min_raw - margin
        y_max_axis = y_max_raw + margin
        major_unit = _nice_major_unit(data_range, target_ticks=6)

        chart1 = wb.add_chart({'type': 'scatter', 'subtype': 'straight'})

        # Si hay balance diario MTM, usarlo como línea principal (continua).
        # La serie por cierre de trade pasa a ser puntos de referencia.
        if has_daily_bal:
            # Ajustar rango Y con el rango del balance diario
            y_min_raw = min(y_min_raw, min(_daily_bals))
            y_max_raw = max(y_max_raw, max(_daily_bals))
            data_range = y_max_raw - y_min_raw if y_max_raw != y_min_raw else max(abs(y_max_raw) * 0.1, 1.0)
            margin     = data_range * 0.03
            y_min_axis = y_min_raw - margin
            y_max_axis = y_max_raw + margin
            major_unit = _nice_major_unit(data_range, target_ticks=6)

        # Series coloreadas por tipo: Estrategia=rojo, B&H=azul
        if has_daily_bal:
            chart1.add_series({
                'name':       'Estrategia',
                'categories': f"='_ChartData'!$I$2:$I${n_daily_points + 1}",
                'values':     f"='_ChartData'!$N$2:$N${n_daily_points + 1}",
                'line':       {'color': '#E53E3E', 'width': 2.0},
                'marker':     {'type': 'none'},
            })
            chart1.add_series({
                'name':       'B&H',
                'categories': f"='_ChartData'!$I$2:$I${n_daily_points + 1}",
                'values':     f"='_ChartData'!$O$2:$O${n_daily_points + 1}",
                'line':       {'color': '#2B6CB0', 'width': 2.0},
                'marker':     {'type': 'none'},
            })

        chart1.set_title({'name': 'Balance', 'name_font': _CHART_TITLE_FONT})
        chart1.set_y_axis({
            'num_format':      '#,##0',
            'min':             y_min_axis,
            'max':             y_max_axis,
            'major_unit':      major_unit,
            **_CHART_AXIS_Y,
        })
        chart1.set_x_axis({
            'num_format':      'MMM yy',
            **_CHART_AXIS_X,
        })
        chart1.set_legend(_CHART_LEGEND)
        chart1.set_plotarea(_CHART_PLOTAREA)
        chart1.set_chartarea(_CHART_CHARTAREA)
        chart1.set_size(_CHART_SIZE)

        ws.insert_chart(BLOCK_ROW, 1, chart1, {'x_offset': 2, 'y_offset': 5})
        CHARTS_HEIGHT_ROWS += 24  # ~430px ≈ 24 filas

    # ── GRÁFICO 2: BALANCE NETO vs BALANCE BRUTO (comparación, solo diario) ──
    if has_daily_bal and n_rows > 0:
        y_min_raw2 = min(_daily_bals + _daily_brutos)
        y_max_raw2 = max(_daily_bals + _daily_brutos)
        data_range2 = y_max_raw2 - y_min_raw2 if y_max_raw2 != y_min_raw2 else max(abs(y_max_raw2) * 0.1, 1.0)
        margin2     = data_range2 * 0.03
        y_min_axis2 = y_min_raw2 - margin2
        y_max_axis2 = y_max_raw2 + margin2
        major_unit2 = _nice_major_unit(data_range2, target_ticks=6)

        chart2 = wb.add_chart({'type': 'scatter', 'subtype': 'straight'})
        chart2.add_series({
            'name':       'Estrategia',
            'categories': f"='_ChartData'!$I$2:$I${n_daily_points + 1}",
            'values':     f"='_ChartData'!$N$2:$N${n_daily_points + 1}",
            'line':       {'color': '#E53E3E', 'width': 2.0},
            'marker':     {'type': 'none'},
        })
        chart2.add_series({
            'name':       'B&H',
            'categories': f"='_ChartData'!$I$2:$I${n_daily_points + 1}",
            'values':     f"='_ChartData'!$O$2:$O${n_daily_points + 1}",
            'line':       {'color': '#2B6CB0', 'width': 2.0},
            'marker':     {'type': 'none'},
        })
        chart2.add_series({
            'name':       'Bruto',
            'categories': f"='_ChartData'!$I$2:$I${n_daily_points + 1}",
            'values':     f"='_ChartData'!$M$2:$M${n_daily_points + 1}",
            'line':       {'color': '#CBD5E0', 'width': 1.5, 'dash_type': 'dash'},
            'marker':     {'type': 'none'},
        })

        chart2.set_title({'name': 'Neto vs Bruto', 'name_font': _CHART_TITLE_FONT})
        chart2.set_y_axis({
            'num_format':      '#,##0',
            'min':             y_min_axis2,
            'max':             y_max_axis2,
            'major_unit':      major_unit2,
            **_CHART_AXIS_Y,
        })
        chart2.set_x_axis({
            'num_format':      'MMM yy',
            **_CHART_AXIS_X,
        })
        chart2.set_legend(_CHART_LEGEND)
        chart2.set_plotarea(_CHART_PLOTAREA)
        chart2.set_chartarea(_CHART_CHARTAREA)
        chart2.set_size(_CHART_SIZE)

        ws.insert_chart(BLOCK_ROW + CHARTS_HEIGHT_ROWS, 1, chart2, {'x_offset': 2, 'y_offset': 5})
        CHARTS_HEIGHT_ROWS += 24

    # ── GRÁFICO 3: ESTRATEGIA vs RENDIMIENTO ACTIVO (scatter XY, siempre diario) ──
    if has_daily_bal and n_rows > 0:
        chart3 = wb.add_chart({'type': 'scatter', 'subtype': 'straight'})

        # Serie 1a — Estrategia ROI% (rojo, solo periodos estrategia)
        chart3.add_series({
            'name':       'Estrategia',
            'categories': f"='_ChartData'!$I$2:$I${n_daily_points + 1}",
            'values':     f"='_ChartData'!$P$2:$P${n_daily_points + 1}",
            'line':       {'color': '#E53E3E', 'width': 2.0},
            'marker':     {'type': 'none'},
        })

        # Serie 1b — B&H ROI% (azul, solo periodos B&H)
        chart3.add_series({
            'name':       'B&H',
            'categories': f"='_ChartData'!$I$2:$I${n_daily_points + 1}",
            'values':     f"='_ChartData'!$Q$2:$Q${n_daily_points + 1}",
            'line':       {'color': '#2B6CB0', 'width': 2.0},
            'marker':     {'type': 'none'},
        })

        # Serie 2 — Rendimiento Activo: precio real, continuo e independiente de trades
        if has_continuous_bh and n_bh_points > 0:
            series_bh = {
                'name':       'Rendimiento Activo',
                'values':     f"='_ChartData'!$H$2:$H${n_bh_points + 1}",
                'categories': f"='_ChartData'!$G$2:$G${n_bh_points + 1}",
                'line':       {'color': '#CBD5E0', 'width': 1.5},
                'marker':     {'type': 'none'},
            }
        else:
            series_bh = {
                'name':       'Rendimiento Activo',
                'values':     f"='_ChartData'!$D$2:$D${n_rows + 1}",
                'categories': f"='_ChartData'!$A$2:$A${n_rows + 1}",
                'line':       {'color': '#CBD5E0', 'width': 1.5},
                'marker':     {'type': 'none'},
            }
        chart3.add_series(series_bh)

        chart3.set_title({'name': 'Estrategia vs Rendimiento Activo', 'name_font': _CHART_TITLE_FONT})
        chart3.set_y_axis({
            'num_format':      '#,##0"%"',
            'crossing':        0,
            **_CHART_AXIS_Y,
        })
        chart3.set_x_axis({
            'num_format':      'MMM yy',
            **_CHART_AXIS_X,
        })
        chart3.set_legend(_CHART_LEGEND)
        chart3.set_plotarea(_CHART_PLOTAREA)
        chart3.set_chartarea(_CHART_CHARTAREA)
        chart3.set_size(_CHART_SIZE)

        ws.insert_chart(BLOCK_ROW + CHARTS_HEIGHT_ROWS, 1, chart3, {'x_offset': 2, 'y_offset': 5})
        CHARTS_HEIGHT_ROWS += 24

    # ── GRÁFICO 4: COMISIONES vs BENEFICIO (siempre diario) ─────────────────
    if has_daily_bal and pnl_neto_idx is not None and n_rows > 0:
        chart4 = wb.add_chart({'type': 'scatter', 'subtype': 'straight'})

        # Serie 1 — PNL Neto acumulado diario (realizado + no realizado)
        chart4.add_series({
            'name':       'Beneficio Neto',
            'values':     f"='_ChartData'!$L$2:$L${n_daily_points + 1}",
            'categories': f"='_ChartData'!$I$2:$I${n_daily_points + 1}",
            'line':       {'color': '#2B6CB0', 'width': 2.0},
            'marker':     {'type': 'none'},
        })

        # Serie 2 — Comisiones acumuladas (siempre por cierre de trade, gris punteado)
        chart4.add_series({
            'name':       'Comisiones',
            'values':     f"='_ChartData'!$E$2:$E${n_rows + 1}",
            'categories': f"='_ChartData'!$A$2:$A${n_rows + 1}",
            'line':       {'color': '#CBD5E0', 'width': 1.5, 'dash_type': 'dash'},
            'marker':     {'type': 'none'},
        })

        chart4.set_title({'name': 'Comisiones vs Beneficio', 'name_font': _CHART_TITLE_FONT})
        chart4.set_y_axis({
            'num_format':      '#,##0',
            **_CHART_AXIS_Y,
        })
        chart4.set_x_axis({
            'num_format':      'MMM yy',
            **_CHART_AXIS_X,
        })
        chart4.set_legend(_CHART_LEGEND)
        chart4.set_plotarea(_CHART_PLOTAREA)
        chart4.set_chartarea(_CHART_CHARTAREA)
        chart4.set_size(_CHART_SIZE)

        ws.insert_chart(BLOCK_ROW + CHARTS_HEIGHT_ROWS, 1, chart4, {'x_offset': 2, 'y_offset': 5})
        CHARTS_HEIGHT_ROWS += 24

    # ── GRÁFICO 5: COMPARATIVA POR PERÍODOS (estrategia vs activo, barras) ──
    if period_data is not None:
        p_labels, p_strat, p_asset = period_data
        n_periods = len(p_labels)

        if n_periods >= 2:
            # Escribir datos en hoja auxiliar oculta _PeriodData
            ws_pd = wb.add_worksheet('_PeriodData')
            ws_pd.hide()
            ws_pd.write_string(0, 0, 'PERIODO')
            ws_pd.write_string(0, 1, 'ESTRATEGIA_%')
            ws_pd.write_string(0, 2, f'{activo_name}_%')

            for pi in range(n_periods):
                ws_pd.write_string(pi + 1, 0, p_labels[pi])
                ws_pd.write_number(pi + 1, 1, p_strat[pi])
                ws_pd.write_number(pi + 1, 2, p_asset[pi])

            chart5 = wb.add_chart({'type': 'column'})

            # Estrategia = gris claro (cerca de blanco), Activo = gris oscuro (cerca de negro)
            _STRAT_COLOR = '#C8CDD3'
            _ASSET_COLOR = '#4A5058'

            chart5.add_series({
                'name':       'Estrategia',
                'categories': f"='_PeriodData'!$A$2:$A${n_periods + 1}",
                'values':     f"='_PeriodData'!$B$2:$B${n_periods + 1}",
                'fill':       {'color': _STRAT_COLOR},
                'border':     {'none': True},
                'gap':        80,
            })
            chart5.add_series({
                'name':       activo_name,
                'categories': f"='_PeriodData'!$A$2:$A${n_periods + 1}",
                'values':     f"='_PeriodData'!$C$2:$C${n_periods + 1}",
                'fill':       {'color': _ASSET_COLOR},
                'border':     {'none': True},
            })

            chart5.set_title({
                'name': f'Estrategia vs {activo_name} — variación % por período',
                'name_font': _CHART_TITLE_FONT,
            })
            chart5.set_y_axis({
                'num_format':      '0.00"%"',
                **_CHART_AXIS_Y,
            })
            chart5.set_x_axis({
                **_CHART_AXIS_X,
            })
            chart5.set_legend(_CHART_LEGEND)
            chart5.set_plotarea(_CHART_PLOTAREA)
            chart5.set_chartarea(_CHART_CHARTAREA)
            chart5.set_size(_CHART_SIZE)

            ws.insert_chart(BLOCK_ROW + CHARTS_HEIGHT_ROWS, 1, chart5, {'x_offset': 2, 'y_offset': 5})

            # ── TABLA RESUMEN junto al chart5 (a la derecha) ────────────────
            _tbl_row = BLOCK_ROW + CHARTS_HEIGHT_ROWS + 1  # un poco abajo del borde superior del chart
            _tbl_col = 15  # columna P (a la derecha del chart de ~14 cols)

            # Calcular victorias por período
            _wins_strat = sum(1 for s, a in zip(p_strat, p_asset) if s > a)
            _wins_asset = sum(1 for s, a in zip(p_strat, p_asset) if a > s)
            _ties = n_periods - _wins_strat - _wins_asset
            # Repartir empates proporcionalmente, o asignar 50/50
            _total_decisive = _wins_strat + _wins_asset if (_wins_strat + _wins_asset) > 0 else 1
            _pct_strat = round(_wins_strat / n_periods * 100, 1)
            _pct_asset = round(_wins_asset / n_periods * 100, 1)
            _pct_ties  = round(100.0 - _pct_strat - _pct_asset, 1)

            # Formatos
            _hdr_fmt = wb.add_format({
                'font_name': 'Calibri', 'font_size': 10, 'bold': True,
                'font_color': '#2D3748', 'bg_color': '#F7FAFC',
                'border': 1, 'border_color': '#E2E8F0',
                'align': 'center', 'valign': 'vcenter',
            })
            _lbl_fmt = wb.add_format({
                'font_name': 'Calibri', 'font_size': 9,
                'font_color': '#718096',
                'border': 1, 'border_color': '#E2E8F0',
                'align': 'left', 'valign': 'vcenter', 'indent': 1,
            })
            _val_fmt = wb.add_format({
                'font_name': 'Calibri', 'font_size': 10, 'bold': True,
                'font_color': '#2D3748',
                'border': 1, 'border_color': '#E2E8F0',
                'align': 'center', 'valign': 'vcenter',
            })
            _pct_fmt = wb.add_format({
                'font_name': 'Calibri', 'font_size': 10, 'bold': True,
                'font_color': '#2D3748',
                'border': 1, 'border_color': '#E2E8F0',
                'align': 'center', 'valign': 'vcenter',
                'num_format': '0.0"%"',
            })

            # Encabezados
            r = _tbl_row
            ws.write_blank(r, _tbl_col, None, _hdr_fmt)
            ws.write_string(r, _tbl_col + 1, 'Estrategia', _hdr_fmt)
            ws.write_string(r, _tbl_col + 2, activo_name, _hdr_fmt)

            # Fila: Períodos ganados
            r += 1
            ws.write_string(r, _tbl_col, 'Períodos ganados', _lbl_fmt)
            ws.write_number(r, _tbl_col + 1, _wins_strat, _val_fmt)
            ws.write_number(r, _tbl_col + 2, _wins_asset, _val_fmt)

            # Fila: Empates (si los hay)
            if _ties > 0:
                r += 1
                ws.write_string(r, _tbl_col, 'Empates', _lbl_fmt)
                ws.write_number(r, _tbl_col + 1, _ties, _val_fmt)
                ws.write_number(r, _tbl_col + 2, _ties, _val_fmt)

            # Fila: % de victoria
            r += 1
            ws.write_string(r, _tbl_col, '% victoria', _lbl_fmt)
            ws.write_number(r, _tbl_col + 1, _pct_strat, _pct_fmt)
            ws.write_number(r, _tbl_col + 2, _pct_asset, _pct_fmt)

            # Fila: Variación acumulada (retorno compuesto)
            import math
            _cum_strat = (math.prod(1 + v / 100 for v in p_strat) - 1) * 100
            _cum_asset = (math.prod(1 + v / 100 for v in p_asset) - 1) * 100
            r += 1
            ws.write_string(r, _tbl_col, 'Var. acumulada', _lbl_fmt)
            ws.write_number(r, _tbl_col + 1, round(_cum_strat, 2), _pct_fmt)
            ws.write_number(r, _tbl_col + 2, round(_cum_asset, 2), _pct_fmt)

            # Ajustar ancho columnas de la tabla
            ws.set_column(_tbl_col, _tbl_col, 16)
            ws.set_column(_tbl_col + 1, _tbl_col + 2, 13)

            CHARTS_HEIGHT_ROWS += 24

    # ── TABLA DE PARÁMETROS ─────────────────────────────────────────────────
    TABLE_COL = 1
    TABLE_ROW = BLOCK_ROW + CHARTS_HEIGHT_ROWS

    _TBL = dict(font_name='Calibri', font_size=11, valign='vcenter', indent=1)

    title_fmt = wb.add_format({**_TBL,
        'bold': True, 'font_color': '#2D3436', 'bg_color': '#EDF2F7',
        'align': 'left',
        'bottom': 2, 'bottom_color': '#A0AEC0',
        'top': 0, 'left': 0, 'right': 0,
    })

    def _lbl_fmt(is_last=False):
        return wb.add_format({**_TBL,
            'bold': False, 'font_color': '#718096', 'bg_color': '#FFFFFF',
            'align': 'left',
            'bottom': 1 if not is_last else 0,
            'bottom_color': '#EDF2F7',
            'top': 0, 'left': 0, 'right': 0,
        })

    def _val_fmt(num_fmt='#,##0.00', is_last=False):
        return wb.add_format({**_TBL,
            'bold': True, 'font_color': '#2D3436', 'bg_color': '#FFFFFF',
            'align': 'right', 'num_format': num_fmt,
            'bottom': 1 if not is_last else 0,
            'bottom_color': '#EDF2F7',
            'top': 0, 'left': 0, 'right': 0,
        })

    items = [
        ("Saldo Usado",    val_saldo,   '#,##0.00 $', False),
        ("Volumen Máx.",   val_volumen, '#,##0.00 $', False),
        ("Apalancamiento", val_apal,    '0.00"x"',    bh_saldo_str is None),
    ]
    if bh_saldo_str is not None:
        _bh_is_all = bh_saldo_str in {"ALL", "TODO", "100%", "100"}
        if not _bh_is_all:
            try:
                _bh_val = float(bh_saldo_str)
                items.append(("Saldo B&H", _bh_val, '#,##0.00 $', True))
            except ValueError:
                _bh_is_all = True
        if _bh_is_all:
            items.append(("Saldo B&H", "Todo el saldo", None, True))

    ws.set_row(TABLE_ROW, 28)
    ws.merge_range(TABLE_ROW, TABLE_COL, TABLE_ROW, TABLE_COL + 1,
                   'Parámetros', title_fmt)

    for i, (label, value, num_fmt, is_last) in enumerate(items):
        r = TABLE_ROW + i + 1
        ws.set_row(r, 24)
        ws.write(r, TABLE_COL, label, _lbl_fmt(is_last))
        if num_fmt is None:
            # Valor de texto (ej. "Todo el saldo")
            _txt_fmt = wb.add_format({**_TBL,
                'bold': True, 'font_color': '#2D3436', 'bg_color': '#FFFFFF',
                'align': 'right', 'italic': True,
                'bottom': 0, 'top': 0, 'left': 0, 'right': 0,
            })
            ws.write_string(r, TABLE_COL + 1, str(value), _txt_fmt)
        else:
            ws.write(r, TABLE_COL + 1, value, _val_fmt(num_fmt, is_last))

    ws.set_column(TABLE_COL,     TABLE_COL,     22)
    ws.set_column(TABLE_COL + 1, TABLE_COL + 1, 18)

    # ── TABLA: ESTRATEGIA vs BUY & HOLD ─────────────────────────────────────
    if regime_mode_breakdown and regime_mode_breakdown.get("has_buy_hold", False):
        rb  = regime_mode_breakdown
        RC  = TABLE_COL + 3   # columna de inicio
        RR  = TABLE_ROW       # fila de inicio

        # ── Formatos específicos de esta tabla ──────────────────────────────
        _rb_title = wb.add_format({
            'font_name': 'Calibri', 'font_size': 11, 'bold': True,
            'font_color': '#2D3436', 'bg_color': '#EDF2F7',
            'align': 'left', 'valign': 'vcenter', 'indent': 1,
            'bottom': 2, 'bottom_color': '#A0AEC0',
        })
        _rb_col_hdr = wb.add_format({
            'font_name': 'Calibri', 'font_size': 9, 'bold': True,
            'font_color': '#FFFFFF', 'bg_color': '#4A5568',
            'align': 'center', 'valign': 'vcenter',
            'border': 1, 'border_color': '#718096',
        })
        _rb_lbl = wb.add_format({
            'font_name': 'Calibri', 'font_size': 9,
            'font_color': '#4A5568', 'bg_color': '#FFFFFF',
            'align': 'left', 'valign': 'vcenter', 'indent': 1,
            'bottom': 1, 'bottom_color': '#EDF2F7',
        })
        _rb_sub = wb.add_format({
            'font_name': 'Calibri', 'font_size': 9, 'bold': True, 'italic': True,
            'font_color': '#718096', 'bg_color': '#F7FAFC',
            'align': 'left', 'valign': 'vcenter', 'indent': 1,
            'top': 1, 'top_color': '#CBD5E0',
            'bottom': 1, 'bottom_color': '#CBD5E0',
        })

        def _rb_val(positive: bool, pct: bool = True, is_last: bool = False):
            """Formato de valor con color según signo."""
            num_fmt = '0.00"%"' if pct else '#,##0.00 $'
            return wb.add_format({
                'font_name': 'Calibri', 'font_size': 10, 'bold': True,
                'font_color': '#38A169' if positive else '#E53E3E',
                'bg_color': '#FFFFFF',
                'align': 'center', 'valign': 'vcenter',
                'num_format': num_fmt,
                'bottom': 1 if not is_last else 0,
                'bottom_color': '#EDF2F7',
            })

        def _rb_neu(pct: bool = True, is_last: bool = False):
            """Formato neutro (sin color) para valores benchmark/count."""
            num_fmt = '0.00"%"' if pct else '0'
            return wb.add_format({
                'font_name': 'Calibri', 'font_size': 10, 'bold': True,
                'font_color': '#2D3436', 'bg_color': '#FFFFFF',
                'align': 'center', 'valign': 'vcenter',
                'num_format': num_fmt,
                'bottom': 1 if not is_last else 0,
                'bottom_color': '#EDF2F7',
            })

        _rb_blank = wb.add_format({
            'bg_color': '#FFFFFF',
            'bottom': 1, 'bottom_color': '#EDF2F7',
        })

        roi_strat       = rb.get("roi_strat", 0.0)
        roi_bh          = rb.get("roi_bh", 0.0)
        roi_total       = rb.get("roi_total", 0.0)
        pnl_strat       = rb.get("pnl_strat", 0.0)
        pnl_bh          = rb.get("pnl_bh", 0.0)
        n_strat         = rb.get("n_strat", 0)
        n_bh            = rb.get("n_bh", 0)
        asset_strat     = rb.get("asset_ret_strat", 0.0)
        asset_bh        = rb.get("asset_ret_bh", 0.0)
        asset_total     = rb.get("asset_ret_total", 0.0)

        # ── Título principal (3 columnas) ────────────────────────────────────
        ws.set_row(RR, 28)
        ws.merge_range(RR, RC, RR, RC + 2, 'Estrategia  vs  Buy & Hold', _rb_title)

        # ── Encabezados de columna ───────────────────────────────────────────
        ws.set_row(RR + 1, 22)
        ws.write_blank(RR + 1, RC,     None,          _rb_col_hdr)
        ws.write(RR + 1, RC + 1, 'Estrategia',        _rb_col_hdr)
        ws.write(RR + 1, RC + 2, 'Buy & Hold',        _rb_col_hdr)

        # ── Fila: ROI (%) ────────────────────────────────────────────────────
        r = RR + 2
        ws.set_row(r, 22)
        ws.write(r, RC,     'ROI (%)',                  _rb_lbl)
        ws.write(r, RC + 1, roi_strat,   _rb_val(roi_strat >= 0,  pct=True))
        ws.write(r, RC + 2, roi_bh,      _rb_val(roi_bh    >= 0,  pct=True))

        # ── Fila: P&L neto ($) ──────────────────────────────────────────────
        r += 1
        ws.set_row(r, 22)
        ws.write(r, RC,     'P&L neto ($)',             _rb_lbl)
        ws.write(r, RC + 1, pnl_strat,  _rb_val(pnl_strat >= 0, pct=False))
        ws.write(r, RC + 2, pnl_bh,     _rb_val(pnl_bh    >= 0, pct=False))

        # ── Fila: Nº operaciones ─────────────────────────────────────────────
        r += 1
        ws.set_row(r, 22)
        ws.write(r, RC,     'Nº operaciones',           _rb_lbl)
        ws.write(r, RC + 1, n_strat,    _rb_neu(pct=False))
        ws.write(r, RC + 2, n_bh,       _rb_neu(pct=False))

        # ── Fila: Benchmark activo (en ese período) ──────────────────────────
        r += 1
        ws.set_row(r, 22)
        ws.write(r, RC,     'Activo (su período)',       _rb_lbl)
        ws.write(r, RC + 1, asset_strat, _rb_val(asset_strat >= 0, pct=True))
        ws.write(r, RC + 2, asset_bh,    _rb_val(asset_bh    >= 0, pct=True))

        # ── Sub-encabezado: Totales ───────────────────────────────────────────
        r += 1
        ws.set_row(r, 20)
        ws.merge_range(r, RC, r, RC + 2, 'Backtest completo', _rb_sub)

        # ── Fila: ROI total combinado ────────────────────────────────────────
        r += 1
        ws.set_row(r, 22)
        ws.write(r, RC,     'ROI total (%)',             _rb_lbl)
        ws.write(r, RC + 1, roi_total,  _rb_val(roi_total >= 0, pct=True))
        ws.write(r, RC + 2, None,       _rb_blank)

        # ── Fila: Activo en todo el backtest ─────────────────────────────────
        r += 1
        ws.set_row(r, 22)
        ws.write(r, RC,     'Activo (backtest completo)', _rb_lbl)
        ws.write(r, RC + 1, asset_total, _rb_val(asset_total >= 0, pct=True, is_last=True))
        ws.write(r, RC + 2, None,        _rb_blank)

        ws.set_column(RC,     RC,     22)
        ws.set_column(RC + 1, RC + 1, 14)
        ws.set_column(RC + 2, RC + 2, 14)

    # ── Duración media por tipo (estrategia vs B&H) ──────────────────────────
    _dur_strat_min: Optional[float] = None
    _dur_bh_min:    Optional[float] = None
    if entry_time_idx is not None and exit_time_idx is not None:
        try:
            _et_col = cols[entry_time_idx]
            _xt_col = cols[exit_time_idx]
            _df_dur = df[[_et_col, _xt_col]].copy()
            _df_dur['_et'] = pd.to_datetime(_df_dur[_et_col], errors='coerce')
            _df_dur['_xt'] = pd.to_datetime(_df_dur[_xt_col], errors='coerce')
            _df_dur['_dur_min'] = (_df_dur['_xt'] - _df_dur['_et']).dt.total_seconds() / 60.0
            if tipo_idx is not None:
                _tipo_col = cols[tipo_idx]
                _df_dur['_tipo'] = df[_tipo_col].values
                _s_dur = _df_dur[_df_dur['_tipo'] == 'S']['_dur_min'].dropna()
                _b_dur = _df_dur[_df_dur['_tipo'] == 'B&H']['_dur_min'].dropna()
                if len(_s_dur) > 0:
                    _dur_strat_min = float(_s_dur.mean())
                if len(_b_dur) > 0:
                    _dur_bh_min = float(_b_dur.mean())
            else:
                _all_dur = _df_dur['_dur_min'].dropna()
                if len(_all_dur) > 0:
                    _dur_strat_min = float(_all_dur.mean())
        except Exception:
            pass

    def _fmt_dur(minutes: Optional[float]) -> str:
        if minutes is None or minutes < 0:
            return '—'
        if minutes < 120:
            return f'{minutes:.0f} min'
        elif minutes < 1440:
            return f'{minutes / 60:.1f} h'
        else:
            return f'{minutes / 1440:.1f} días'

    # ── TABLA: RESUMEN EJECUTIVO ─────────────────────────────────────────────
    _SUM_COL = TABLE_COL + 7   # cols 8-9 (derecha de la tabla B&H + gap)
    _SUM_ROW = TABLE_ROW

    # ── Recoger valores ──────────────────────────────────────────────────────
    _sum_asset_growth: Optional[float] = None
    _sum_roi_strat:    Optional[float] = None
    _sum_roi_bh:       Optional[float] = None
    _sum_roi_total:    Optional[float] = None
    _sum_n_strat:      Optional[int]   = None
    _sum_n_bh:         Optional[int]   = None

    if regime_mode_breakdown:
        _sum_asset_growth = regime_mode_breakdown.get("asset_ret_total")
        _sum_roi_strat    = regime_mode_breakdown.get("roi_strat")
        _sum_roi_bh       = regime_mode_breakdown.get("roi_bh")
        _sum_roi_total    = regime_mode_breakdown.get("roi_total")
        _sum_n_strat      = regime_mode_breakdown.get("n_strat")
        _sum_n_bh         = regime_mode_breakdown.get("n_bh")

    # Fallback asset_growth desde precio cacheado
    if _sum_asset_growth is None and price_series and len(price_series) >= 2:
        _p0, _p1 = price_series[0], price_series[-1]
        if _p0 > 0:
            _sum_asset_growth = (_p1 / _p0 - 1.0) * 100.0

    # Fallback ROI estrategia desde df
    if _sum_roi_strat is None and n_rows > 0 and balance_idx is not None and pnl_neto_idx is not None:
        try:
            _f_pnl = float(df.iloc[0, pnl_neto_idx])
            _f_bal = float(df.iloc[0, balance_idx])
            _l_bal = float(df.iloc[-1, balance_idx])
            _s0    = _f_bal - _f_pnl
            if _s0 > 0:
                _sum_roi_strat = (_l_bal / _s0 - 1.0) * 100.0
                _sum_roi_total = _sum_roi_strat
        except Exception:
            pass

    # ROI total fallback: si no vino de breakdown, suma parciales
    if _sum_roi_total is None:
        _a = _sum_roi_strat or 0.0
        _b = _sum_roi_bh    or 0.0
        _sum_roi_total = _a + _b if (_a != 0.0 or _b != 0.0) else None

    # Fallback conteo trades
    if _sum_n_strat is None:
        _sum_n_strat = n_rows
    if _sum_n_bh is None:
        _sum_n_bh = 0

    # ── Paleta de colores de la tabla ────────────────────────────────────────
    _C_HDR_BG   = '#1A202C'   # cabecera oscura
    _C_HDR_TXT  = '#FFFFFF'
    _C_SUB_BG   = '#2D3748'   # sub-cabecera (ROI / Trades)
    _C_SUB_TXT  = '#E2E8F0'
    _C_LBL_BG   = '#F7FAFC'
    _C_LBL_TXT  = '#4A5568'
    _C_TOT_BG   = '#EBF8FF'   # fila total, fondo azul muy suave
    _C_TOT_TXT  = '#2B6CB0'
    _C_VAL_BG   = '#FFFFFF'
    _C_GREEN    = '#276749'
    _C_RED      = '#C53030'
    _C_BORDER   = '#CBD5E0'

    def _sf(bg=_C_VAL_BG, fc='#2D3436', bold=False, sz=10,
            num_fmt=None, align='right', bottom=1, italic=False, indent=0):
        p = {
            'font_name': 'Calibri', 'font_size': sz, 'bold': bold, 'italic': italic,
            'font_color': fc, 'bg_color': bg,
            'align': align, 'valign': 'vcenter', 'indent': indent,
            'border': 0,
            'bottom': bottom, 'bottom_color': _C_BORDER,
            'left': 1, 'left_color': _C_BORDER,
            'right': 1, 'right_color': _C_BORDER,
        }
        if num_fmt:
            p['num_format'] = num_fmt
        return wb.add_format(p)

    # Formatos reutilizables
    _sf_hdr      = _sf(bg=_C_HDR_BG,  fc=_C_HDR_TXT,  bold=True, sz=11,
                       align='center', bottom=2, indent=1)
    _sf_sub      = _sf(bg=_C_SUB_BG,  fc=_C_SUB_TXT,  bold=True, sz=9,
                       align='left',  bottom=1, indent=1)
    _sf_sub_val  = _sf(bg=_C_SUB_BG,  fc=_C_SUB_TXT,  bold=True, sz=9,
                       align='center', bottom=1)
    _sf_lbl      = _sf(bg=_C_LBL_BG,  fc=_C_LBL_TXT,  bold=False, sz=9,
                       align='left',  bottom=1, indent=2)
    _sf_lbl_last = _sf(bg=_C_LBL_BG,  fc=_C_LBL_TXT,  bold=False, sz=9,
                       align='left',  bottom=2, indent=2)
    _sf_tot_lbl  = _sf(bg=_C_TOT_BG,  fc=_C_TOT_TXT,  bold=True, sz=10,
                       align='left',  bottom=2, indent=2)
    _sf_na       = _sf(bg=_C_VAL_BG,  fc='#A0AEC0',   bold=False, sz=9,
                       align='center', bottom=1, italic=True)
    _sf_na_last  = _sf(bg=_C_VAL_BG,  fc='#A0AEC0',   bold=False, sz=9,
                       align='center', bottom=2, italic=True)
    _sf_na_tot   = _sf(bg=_C_TOT_BG,  fc='#A0AEC0',   bold=False, sz=9,
                       align='center', bottom=2, italic=True)

    def _sf_pct(positive=True, is_last=False, is_total=False, is_neutral=False):
        if is_neutral:
            fc, bg = '#2D3748', _C_LBL_BG
        elif is_total:
            fc, bg = _C_TOT_TXT, _C_TOT_BG
        else:
            fc  = _C_GREEN if positive else _C_RED
            bg  = _C_VAL_BG
        return _sf(bg=bg, fc=fc, bold=True, sz=10,
                   num_fmt='0.00"%"', align='center',
                   bottom=2 if is_last else 1)

    def _sf_int(is_last=False, is_total=False):
        fc = _C_TOT_TXT if is_total else '#2D3748'
        bg = _C_TOT_BG  if is_total else _C_VAL_BG
        return _sf(bg=bg, fc=fc, bold=True, sz=10,
                   num_fmt='0', align='center',
                   bottom=2 if is_last else 1)

    # ── Construcción de la tabla ─────────────────────────────────────────────
    _r = _SUM_ROW

    # Encabezado principal (merge 2 cols)
    ws.set_row(_r, 30)
    ws.merge_range(_r, _SUM_COL, _r, _SUM_COL + 1, 'RESUMEN EJECUTIVO', _sf_hdr)

    # Sub-sección: RENDIMIENTO
    _r += 1
    ws.set_row(_r, 20)
    ws.write(_r, _SUM_COL,     'RENDIMIENTO',  _sf_sub)
    ws.write_blank(_r, _SUM_COL + 1, None,     _sf_sub_val)

    # Fila: Crecimiento del activo (benchmark neutro)
    _r += 1
    ws.set_row(_r, 22)
    ws.write(_r, _SUM_COL, 'Crecim. activo', _sf_lbl)
    if _sum_asset_growth is not None:
        ws.write_number(_r, _SUM_COL + 1, _sum_asset_growth,
                        _sf_pct(positive=(_sum_asset_growth >= 0), is_neutral=True))
    else:
        ws.write_string(_r, _SUM_COL + 1, '—', _sf_na)

    # Fila: ROI estrategia
    _r += 1
    ws.set_row(_r, 22)
    ws.write(_r, _SUM_COL, 'ROI estrategia', _sf_lbl)
    if _sum_roi_strat is not None:
        ws.write_number(_r, _SUM_COL + 1, _sum_roi_strat,
                        _sf_pct(positive=(_sum_roi_strat >= 0)))
    else:
        ws.write_string(_r, _SUM_COL + 1, '—', _sf_na)

    # Fila: ROI B&H
    _r += 1
    ws.set_row(_r, 22)
    ws.write(_r, _SUM_COL, 'ROI B&H', _sf_lbl)
    if _sum_roi_bh is not None:
        ws.write_number(_r, _SUM_COL + 1, _sum_roi_bh,
                        _sf_pct(positive=(_sum_roi_bh >= 0)))
    else:
        ws.write_string(_r, _SUM_COL + 1, '—', _sf_na)

    # Fila: ROI TOTAL (destacada)
    _r += 1
    ws.set_row(_r, 24)
    ws.write(_r, _SUM_COL, 'ROI TOTAL', _sf_tot_lbl)
    if _sum_roi_total is not None:
        ws.write_number(_r, _SUM_COL + 1, _sum_roi_total,
                        _sf_pct(positive=(_sum_roi_total >= 0), is_total=True, is_last=True))
    else:
        ws.write_string(_r, _SUM_COL + 1, '—', _sf_na_tot)

    # Sub-sección: OPERACIONES
    _r += 1
    ws.set_row(_r, 20)
    ws.write(_r, _SUM_COL,     'OPERACIONES',  _sf_sub)
    ws.write_blank(_r, _SUM_COL + 1, None,     _sf_sub_val)

    # Fila: Trades estrategia
    _r += 1
    ws.set_row(_r, 22)
    ws.write(_r, _SUM_COL, 'Trades estrategia', _sf_lbl)
    ws.write_number(_r, _SUM_COL + 1, _sum_n_strat, _sf_int())

    # Fila: Trades B&H
    _r += 1
    ws.set_row(_r, 22)
    ws.write(_r, _SUM_COL, 'Trades B&H', _sf_lbl)
    ws.write_number(_r, _SUM_COL + 1, _sum_n_bh, _sf_int())

    # Fila: Duración media estrategia
    _sf_dur_val = _sf(bg=_C_VAL_BG, fc='#2D3748', bold=True, sz=10,
                      align='center', bottom=1)
    _sf_dur_val_last = _sf(bg=_C_VAL_BG, fc='#2D3748', bold=True, sz=10,
                           align='center', bottom=2)
    _r += 1
    ws.set_row(_r, 22)
    ws.write(_r, _SUM_COL, 'Dur. media estrat.', _sf_lbl)
    ws.write_string(_r, _SUM_COL + 1, _fmt_dur(_dur_strat_min), _sf_dur_val)

    # Fila: Duración media B&H
    _r += 1
    ws.set_row(_r, 22)
    ws.write(_r, _SUM_COL, 'Dur. media B&H', _sf_lbl_last)
    ws.write_string(_r, _SUM_COL + 1, _fmt_dur(_dur_bh_min), _sf_dur_val_last)

    ws.set_column(_SUM_COL,     _SUM_COL,     20)
    ws.set_column(_SUM_COL + 1, _SUM_COL + 1, 13)

    wb.close()


# ==============================================================================
# FALLBACK OPENPYXL
# ==============================================================================

def _escribir_trades_openpyxl_fallback(
    filepath: str,
    df: pd.DataFrame,
    val_saldo: float = 0,
    val_volumen: float = 0,
    val_apal: float = 0,
):
    from openpyxl.formatting.rule import CellIsRule

    df.to_excel(filepath, index=False, sheet_name="Trades", engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column
    max_row = ws.max_row

    border   = _make_border_op(COLORS["border_color"])
    fill_hdr = PatternFill("solid", fgColor=COLORS["header_bg_metrics"])
    fill_alt = PatternFill("solid", fgColor=COLORS["row_alt"])
    fill_w   = PatternFill("solid", fgColor="FFFFFF")
    font_hdr = Font(name=FONT_TITLE, size=11, bold=True, color="FFFFFF")
    font_b   = Font(name=FONT_BODY,  size=10, color=COLORS["text_dark"])
    align_c  = Alignment(horizontal='center', vertical='center')
    align_cw = Alignment(horizontal='center', vertical='center', wrap_text=True)

    col_headers = {c: str(ws.cell(1, c).value or "").upper() for c in range(1, max_col + 1)}
    col_fmt = {}
    pnl_col = None
    for c, hdr in col_headers.items():
        if "TIME" in hdr or "DATE" in hdr:
            col_fmt[c] = "DD/MM/YY HH:MM"
        elif "PRICE" in hdr:
            col_fmt[c] = "#,##0.00"
        elif "PNL" in hdr or "BALANCE" in hdr or "COMISIONES" in hdr or "VOLUMEN" in hdr:
            col_fmt[c] = "#,##0.00"
        else:
            col_fmt[c] = "General"
        if "PNL" in hdr and "NETO" in hdr:
            pnl_col = c

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font_hdr; cell.alignment = align_cw
        cell.fill = fill_hdr; cell.border = border
    ws.row_dimensions[1].height = 28

    for r in range(2, max_row + 1):
        rf = fill_alt if r % 2 == 0 else fill_w
        ws.row_dimensions[r].height = 18
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_b; cell.alignment = align_c
            cell.border = border; cell.fill = rf
            cell.number_format = col_fmt.get(c, "General")

    if pnl_col:
        col_letter = get_column_letter(pnl_col)
        rng = f"{col_letter}2:{col_letter}{max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='greaterThan', formula=['0'],
            fill=PatternFill("solid", fgColor=COLORS["success_bg"]),
            font=Font(name=FONT_BODY, size=10, color=COLORS["accent_green"], bold=True)
        ))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='lessThan', formula=['0'],
            fill=PatternFill("solid", fgColor=COLORS["danger_bg"]),
            font=Font(name=FONT_BODY, size=10, color=COLORS["accent_red"], bold=True)
        ))

    _col_widths_fast_op(ws, max_col, max_row)
    ws.freeze_panes = ws.cell(row=2, column=1)

    # Gráfico openpyxl con eje Y escalado
    col_map = {str(ws.cell(1, c).value or "").upper().strip(): c for c in range(1, max_col + 1)}
    if "BALANCE" in col_map:
        from openpyxl.chart import LineChart, Reference
        bal_col_idx = col_map["BALANCE"]
        bal_data = [ws.cell(r, bal_col_idx).value for r in range(2, max_row + 1)
                    if isinstance(ws.cell(r, bal_col_idx).value, (int, float))]
        if bal_data:
            data_range = max(bal_data) - min(bal_data) or 1
            margin = data_range * 0.03
            chart2 = LineChart()
            chart2.title  = "EVOLUCIÓN DEL BALANCE"
            chart2.legend = None
            chart2.y_axis.numFmt = '#,##0.00'
            chart2.y_axis.scaling.min = min(bal_data) - margin
            chart2.y_axis.scaling.max = max(bal_data) + margin
            chart2.y_axis.majorGridlines = None
            chart2.x_axis.majorGridlines = None
            data = Reference(ws, min_col=bal_col_idx, min_row=1, max_row=max_row)
            chart2.add_data(data, titles_from_data=True)
            s1 = chart2.series[0]
            s1.graphicalProperties.line.solidFill = "3A86FF"
            s1.graphicalProperties.line.width = 22860
            s1.smooth = True
            s1.marker.symbol = "none"
            chart2.height = 14; chart2.width = 44  # Ampliado (ancho 10 eq)
            ws.add_chart(chart2, f"B{max_row + 3}")

    # Tabla fallback debajo del gráfico (B)
    TABLE_COL_OP = 2   # openpyxl 1-indexed = B
    TABLE_ROW_OP = max_row + 33  # Justo debajo del gráfico (que es alto)
    for i, (label, value) in enumerate([("SALDO", val_saldo), ("VOLUMEN", val_volumen), ("APALANCAMIENTO", val_apal)]):
        r = TABLE_ROW_OP + i
        ws.cell(r, TABLE_COL_OP).value = label
        ws.cell(r, TABLE_COL_OP + 1).value = value

    wb.save(filepath)

# ==============================================================================
# FUNCIÓN PRINCIPAL (CSV → DASHBOARD RESUMEN)
# ==============================================================================

def _load_price_with_timestamps(
    activo: str,
    datos_dir: str,
    fecha_inicio: Optional[str],
    fecha_fin: Optional[str],
    n_points: int = 200,
) -> Tuple[Optional[List[float]], Optional[List]]:
    """Carga precios + timestamps del activo para gráfica B&H continua.
    
    Usa resampling diario para reducir millones de velas 1m a ~1500 filas,
    y luego selecciona n_points equidistantes. Ultra rápido (~200ms).
    """
    try:
        import glob as _glob
        import polars as _pl

        activo_up = str(activo).strip().upper()
        base_dir  = datos_dir or "datos"

        candidates = [
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.feather"),
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.fthr"),
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.parquet"),
            os.path.join(base_dir, f"{activo_up.lower()}_ohlcv_1m.feather"),
            os.path.join(base_dir, f"{activo_up.lower()}_ohlcv_1m.parquet"),
        ]

        filepath = next((p for p in candidates if os.path.exists(p)), None)
        if filepath is None:
            for pat in [f"{activo_up}_ohlcv_1m.*", f"{activo_up.lower()}_ohlcv_1m.*"]:
                hits = _glob.glob(os.path.join(base_dir, pat))
                if hits:
                    filepath = hits[0]
                    break
        if filepath is None:
            return None, None

        ext = os.path.splitext(filepath)[1].lower()

        # Lazy scan para no cargar todo en memoria
        if ext in (".feather", ".fthr"):
            # Feather no soporta scan_ipc bien con compresión → leer solo columnas necesarias
            df_raw = _pl.read_ipc(filepath, columns=["timestamp", "close"], memory_map=False)
        elif ext in (".parquet", ".pq"):
            df_raw = _pl.scan_parquet(filepath).select(["timestamp", "close"]).collect()
        else:
            return None, None

        # Normalizar timestamp a Datetime[us] naive
        ts_col = df_raw.get_column("timestamp")
        if str(ts_col.dtype).startswith("Datetime"):
            try:
                df_raw = df_raw.with_columns(
                    _pl.col("timestamp").cast(_pl.Datetime("us")).alias("timestamp")
                )
            except Exception:
                try:
                    df_raw = df_raw.with_columns(
                        _pl.col("timestamp").dt.replace_time_zone(None).cast(_pl.Datetime("us")).alias("timestamp")
                    )
                except Exception:
                    pass

        # Filtrar por rango de fechas
        if fecha_inicio:
            try:
                dt_start = _pl.Series([fecha_inicio]).str.to_datetime(format="%Y-%m-%d", strict=False).item()
                df_raw = df_raw.filter(_pl.col("timestamp") >= dt_start)
            except Exception:
                pass
        if fecha_fin:
            try:
                dt_end = _pl.Series([fecha_fin]).str.to_datetime(format="%Y-%m-%d", strict=False).item()
                df_raw = df_raw.filter(_pl.col("timestamp") <= dt_end)
            except Exception:
                pass

        if df_raw.height < 2:
            return None, None

        # Resamplear a 1D para reducir de ~2M filas a ~1500 filas
        df_daily = (
            df_raw
            .sort("timestamp")
            .group_by_dynamic("timestamp", every="1d")
            .agg(_pl.col("close").last())
            .drop_nulls()
        )

        if df_daily.height < 2:
            return None, None

        prices_all = df_daily.get_column("close").to_list()
        ts_all     = df_daily.get_column("timestamp").to_list()
        n_src = len(prices_all)

        if n_points <= 1:
            return [prices_all[0]], [ts_all[0]]

        # Seleccionar n_points equidistantes
        if n_src <= n_points:
            return prices_all, ts_all

        indices     = [int(round(i * (n_src - 1) / (n_points - 1))) for i in range(n_points)]
        resampled_p = [prices_all[i] for i in indices]
        resampled_t = [ts_all[i]     for i in indices]
        return resampled_p, resampled_t

    except Exception as _e:
        logger.debug(f"_load_price_with_timestamps: {_e}")
        return None, None

def _load_price_series(
    activo: str,
    datos_dir: str,
    fecha_inicio: Optional[str],
    fecha_fin: Optional[str],
    formato_datos: str,
    n_points: int,
) -> Optional[List[float]]:
    """
    Carga precios de cierre del activo desde el archivo de datos OHLCV y los
    remuestrea a exactamente n_points valores con espaciado uniforme.

    Retorna None si el archivo no existe o hay algún error.
    """
    try:
        import glob as _glob
        import polars as _pl

        activo_up = str(activo).strip().upper()
        base_dir  = datos_dir or "datos"

        # Candidatos en orden de preferencia
        candidates = [
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.feather"),
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.fthr"),
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.parquet"),
            os.path.join(base_dir, f"{activo_up}_ohlcv_1m.pq"),
            os.path.join(base_dir, f"{activo_up.lower()}_ohlcv_1m.feather"),
            os.path.join(base_dir, f"{activo_up.lower()}_ohlcv_1m.parquet"),
        ]

        filepath = next((p for p in candidates if os.path.exists(p)), None)
        if filepath is None:
            # Búsqueda glob como fallback
            for pat in [f"{activo_up}_ohlcv_1m.*", f"{activo_up.lower()}_ohlcv_1m.*"]:
                hits = _glob.glob(os.path.join(base_dir, pat))
                if hits:
                    filepath = hits[0]
                    break

        if filepath is None:
            return None

        ext = os.path.splitext(filepath)[1].lower()
        cols = ["timestamp", "close"]
        if ext in (".feather", ".fthr"):
            df_raw = _pl.read_ipc(filepath, columns=cols, memory_map=False)
        elif ext in (".parquet", ".pq"):
            df_raw = _pl.read_parquet(filepath, columns=cols)
        else:
            return None

        # Normalizar timestamp a datetime[μs] naive para evitar errores de comparación
        ts_col = df_raw.get_column("timestamp")
        if ts_col.dtype == _pl.Datetime("ns", "UTC") or str(ts_col.dtype).startswith("Datetime"):
            try:
                df_raw = df_raw.with_columns(
                    _pl.col("timestamp").cast(_pl.Datetime("us")).alias("timestamp")
                )
            except Exception:
                try:
                    df_raw = df_raw.with_columns(
                        _pl.col("timestamp").dt.replace_time_zone(None).cast(_pl.Datetime("us")).alias("timestamp")
                    )
                except Exception:
                    pass

        # Filtrado por rango de fechas
        if fecha_inicio:
            try:
                dt_start = _pl.Series([fecha_inicio]).str.to_datetime(format="%Y-%m-%d", strict=False).item()
                df_raw = df_raw.filter(_pl.col("timestamp") >= dt_start)
            except Exception:
                pass
        if fecha_fin:
            try:
                dt_end = _pl.Series([fecha_fin]).str.to_datetime(format="%Y-%m-%d", strict=False).item()
                df_raw = df_raw.filter(_pl.col("timestamp") <= dt_end)
            except Exception:
                pass

        if df_raw.height < 2:
            return None

        # Resamplear a 1D para reducir de ~2M filas a ~1500 filas (ultra rápido)
        df_daily = (
            df_raw
            .sort("timestamp")
            .group_by_dynamic("timestamp", every="1d")
            .agg(_pl.col("close").last())
            .drop_nulls()
        )

        if df_daily.height < 2:
            return None

        prices = df_daily.get_column("close").to_list()
        n_src  = len(prices)

        if n_points <= 1:
            return [prices[0]]

        if n_src <= n_points:
            return prices

        # Remuestreo: selección de índices equidistantes
        indices  = [int(round(i * (n_src - 1) / (n_points - 1))) for i in range(n_points)]
        resampled = [prices[i] for i in indices]
        return resampled

    except Exception as _e:
        logger.debug(f"_load_price_series: {_e}")
        return None


def convertir_resumen_csv_a_excel(
    *,
    csv_path: str,
    strategy_name: str,
    activo: Optional[str] = None,
    timeframe: Optional[str] = None,
    saldo_inicial: float = 300.0,
    excel_path: Optional[str] = None,
    output_dir: Optional[str] = None,
    # Datos de precio para la gráfica
    datos_dir: str = "datos",
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    formato_datos: str = "feather",
) -> str:
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"No existe el CSV: {csv_path}")

    if output_dir is None and excel_path:
        output_dir = os.path.dirname(str(excel_path)) or None

    activo    = activo or "UNKNOWN"
    timeframe = timeframe or "UNKNOWN"

    df = pd.read_csv(csv_path)
    df, known_param_names = _normalizar_nombres(df, strategy_name)
    df_final, cols_metrics, cols_params = _organizar_y_filtrar_columnas(df, known_param_names)
    df_final = _ordenar_filas(df_final)

    final_excel_path = _generar_nombre_archivo(csv_path, output_dir, str(activo), strategy_name, str(timeframe))
    os.makedirs(os.path.dirname(final_excel_path) or ".", exist_ok=True)

    with pd.ExcelWriter(final_excel_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, startrow=1)

    _aplicar_estilo_avanzado(
        final_excel_path, df_final, cols_metrics, cols_params, saldo_inicial,
    )

    path_to_return = final_excel_path
    if excel_path:
        import shutil
        try:
            if os.path.abspath(excel_path) != os.path.abspath(final_excel_path):
                shutil.move(final_excel_path, excel_path)
                path_to_return = excel_path
        except Exception:
            pass

    try:
        if os.path.exists(csv_path):
            os.remove(csv_path)
    except Exception:
        pass

    return path_to_return


# ==============================================================================
# LÓGICA DE PROCESAMIENTO
# ==============================================================================

def _normalizar_nombres(df: pd.DataFrame, strategy_name: str) -> tuple:
    """
    Normaliza nombres de columnas y detecta cuáles son parámetros de estrategia.

    DETECCIÓN ROBUSTA DE PARÁMETROS:
    ================================
    En el CSV intermedio, los parámetros de la estrategia se guardan con el
    prefijo 'PARAM_' (ej: PARAM_OLS_WINDOW, PARAM_EXIT_SL_PCT). Usamos
    este prefijo como FUENTE DE VERDAD para saber qué es un parámetro,
    en lugar de depender de heurísticas de keywords que fallan con nuevas
    estrategias.

    Returns:
        tuple: (df_normalizado, known_param_names) donde known_param_names
               es un set con los nombres limpios de las columnas que
               originalmente tenían prefijo PARAM_.
    """
    df.columns = [str(c).upper().strip() for c in df.columns]

    # ── FASE 1: Registrar columnas que vienen con prefijo PARAM_ ──────────
    # Estas son DEFINITIVAMENTE parámetros de la estrategia.
    # Guardamos el nombre limpio (sin prefijo) para usarlo después.
    _PARAM_PREFIXES = ("PARAM_", "PARAMS_", "ESTRATEGIA_PARAMS_", "STRATEGY_PARAMS_")
    known_param_names: set = set()
    for col in df.columns:
        for pfx in _PARAM_PREFIXES:
            if col.startswith(pfx):
                clean_name = col[len(pfx):]
                # Excluir params internos (__activo, __saldo, etc.)
                if not clean_name.startswith("_"):
                    known_param_names.add(clean_name)
                break

    # ── FASE 2: Renombrado estándar de métricas y columnas conocidas ──────
    rename_map = {
        "STRATEGY": "ESTRATEGIA", "SHARPE_RATIO": "SHARPE", "TRADES_POR_DIA": "TRADES_DIA",
        "N_TRADES": "TOTAL_TRADES", "NUM_TRADES": "TOTAL_TRADES", "COUNT_TRADES": "TOTAL_TRADES",
        "AVG_TRADE_DURATION": "AVG_TRADE", "DURATION_MEAN_MIN": "AVG_TRADE",
        "RETORNO_PROMEDIO": "AVG_TRADE",
        "WIN_RATE_PCT": "WINRATE_PCT", "WINRATE": "WINRATE_PCT",
        "PORC_GANADORAS": "WINRATE_PCT", "WIN_RATE": "WINRATE_PCT",
        "RACHA_GANADORA": "WIN_STREAK", "RACHA_PERDEDORA": "LOSS_STREAK",
        "MAX_DRAWDOWN_PCT": "MAX_DD_PCT", "MAX_DRAWDOWN": "MAX_DD_PCT",
        "DRAWDOWN": "MAX_DD_PCT", "DD": "MAX_DD_PCT", "DD_PCT": "MAX_DD_PCT", "MAX_DD": "MAX_DD_PCT",
        "RETURN_PCT": "ROI_PCT", "ROI": "ROI_PCT",
        "COUNT_LONGS": "LONG", "N_LONGS": "LONG", "LONGS": "LONG", "NUM_LONGS": "LONG",
        "N_TRADES_LONG": "LONG", "TRADES_LONG": "LONG",
        "COUNT_SHORTS": "SHORT", "N_SHORTS": "SHORT", "SHORTS": "SHORT", "NUM_SHORTS": "SHORT",
        "N_TRADES_SHORT": "SHORT", "TRADES_SHORT": "SHORT",
        "EXIT_SL_PCT": "SL", "P_SL": "SL", "SL_PCT": "SL",
        "EXIT_TP_PCT": "TP", "P_TP": "TP", "TP_PCT": "TP",
        "EXIT_TRAIL_ACT_PCT": "ACT", "TRAIL_ACT": "ACT",
        "EXIT_TRAIL_DIST_PCT": "DIST", "TRAIL_DIST": "DIST", "DISTANCE": "DIST",
        # --- Mapeos para las métricas del nuevo resumen ---
        "NET_PNL": "PNL_NETO",
        "SALDO_BRUTO": "SALDO_SIN_COMISIONES",
        "COMISIONES": "COMISIONES_TOTAL",
        "COMISION_TOTAL": "COMISIONES_TOTAL",
    }

    for old, new in rename_map.items():
        if old in df.columns and new not in df.columns:
            df.rename(columns={old: new}, inplace=True)
        elif old in df.columns and new in df.columns:
            df.drop(columns=[old], inplace=True)

    if "ESTRATEGIA" not in df.columns:
        df.insert(0, "ESTRATEGIA", strategy_name.upper())
    if "TRIAL" not in df.columns and df.index.name != "TRIAL":
        df.insert(0, "TRIAL", range(len(df)))

    # ── FASE 3: Limpiar prefijos y normalizar nombres ─────────────────────
    # Mantenemos un mapeo old_name → clean_name para actualizar known_param_names
    def _canon_metric(col_name: str) -> str:
        return METRIC_ALIASES_TO_CANONICAL.get(col_name, col_name)

    new_cols = []
    updated_param_names: set = set()
    for col in df.columns:
        if col in METRICS_ORDER or col in ID_COLS or col in {"SL", "TP", "ACT", "DIST"}:
            # SL/TP/ACT/DIST son parámetros de salida conocidos, marcarlos
            if col in {"SL", "TP", "ACT", "DIST"}:
                updated_param_names.add(col)
            new_cols.append(col)
            continue

        clean = col
        was_param = False
        for p in PREFIXES_TO_CLEAN:
            if clean.startswith(p):
                # Si tenía prefijo de parámetro, marcarlo
                if p in ("PARAM_", "PARAMS_", "ESTRATEGIA_PARAMS_", "STRATEGY_PARAMS_"):
                    was_param = True
                clean = clean[len(p):]

        # Excluir params internos del sistema (__, doble guión bajo)
        if clean.startswith("_"):
            # Son params internos (__activo, __saldo, etc.), no mostrar
            new_cols.append(clean)
            continue

        # Aplicar rename_map DESPUÉS de limpiar prefijos para atrapar
        # casos como PARAM_EXIT_SL_PCT → (strip PARAM_) → EXIT_SL_PCT → SL
        if clean in rename_map:
            clean = rename_map[clean]
        clean = clean.replace("_PCT", "%").replace("PERCENTAGE", "%")
        clean = _canon_metric(clean)

        # Si el nombre original (sin limpiar prefijo) estaba en known_param_names,
        # o si fue detectado por tener prefijo PARAM_, marcar el nombre limpio
        original_without_prefix = col
        for pfx in _PARAM_PREFIXES:
            if original_without_prefix.startswith(pfx):
                original_without_prefix = original_without_prefix[len(pfx):]
                break
        if was_param or original_without_prefix in known_param_names:
            updated_param_names.add(clean)

        new_cols.append(clean)

    df.columns = new_cols

    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated()]

    return df, updated_param_names


def _organizar_y_filtrar_columnas(df: pd.DataFrame, known_param_names: set = None):
    """
    Organiza columnas en: ID | MÉTRICAS | PARÁMETROS.

    DETECCIÓN ROBUSTA:
    ==================
    1. PRIORIDAD 1 — known_param_names: Columnas que vinieron con prefijo PARAM_
       en el CSV original. Estas son DEFINITIVAMENTE parámetros, sin importar
       su nombre. Esto garantiza que cualquier parámetro de cualquier estrategia
       nueva sea detectado correctamente.

    2. PRIORIDAD 2 — Fallback heurístico: Para columnas que no están en
       known_param_names (ej: datos legacy sin prefijo), se usa la heurística
       de keywords existente como respaldo.
    """
    if known_param_names is None:
        known_param_names = set()

    cols = list(df.columns)
    current_ids = [c for c in ID_COLS if c in cols]

    current_metrics = []
    for m in ALL_METRICS_ORDER:
        for t in [m, m.replace("_PCT", "%")]:
            if t in cols:
                current_metrics.append(t)
                break
    current_metrics = list(dict.fromkeys(current_metrics))

    excluded = set(current_ids + current_metrics) | ALL_KNOWN_METRICS
    candidates = [c for c in cols if c not in excluded]

    exit_cols = {"SL", "TP", "ACT", "DIST"}

    # Determinar si es TRAILING o FIXED
    is_trailing = False
    exit_col_name = next((c for c in df.columns if str(c).upper() in ["EXIT_TYPE", "PARAM_EXIT_TYPE", "PARAM_EXIT_TYPE"]), None)
    if exit_col_name:
        is_trailing = df[exit_col_name].astype(str).str.contains("TRAIL", case=False).any()
    else:
        for t_col in ["ACT", "DIST"]:
            if t_col in df.columns:
                try:
                    if (pd.to_numeric(df[t_col], errors='coerce').fillna(0) > 0).any():
                        is_trailing = True; break
                except Exception:
                    pass

    # ── Clasificación de parámetros con doble vía ─────────────────────────
    very_bad = {"PROFIT", "WIN", "SALDO", "BALANCE", "DRAWDOWN", "DD",
                "ROI", "RETORNO", "NUM_", "COUNT", "TRADES", "RESULT", "METRIC"}
    exceptions = {"STOP", "SL", "TP", "TRAIL", "TIME", "PERIOD", "LEN", "FAST",
                  "SLOW", "SIGNAL", "LIMIT", "THRESHOLD", "SIGMA", "OFFSET", "ATR", "ACT", "DIST"}

    current_params = []
    for c in candidates:
        # Saltar columnas vacías o internas del sistema
        if df[c].astype(str).str.strip().eq("").all() or c.startswith("_"):
            continue
        if c in EXCLUDED_PARAMS or c.replace("%", "_PCT") in EXCLUDED_PARAMS:
            continue

        # ── VÍA 1: ¿Está en known_param_names? → ES PARÁMETRO (definitivo) ──
        if c in known_param_names:
            # Respetar lógica de trailing para ACT/DIST
            if c in {"ACT", "DIST"} and not is_trailing:
                continue
            current_params.append(c)
            continue

        # ── VÍA 2: Exit cols conocidas ───────────────────────────────────────
        if c in exit_cols:
            if c in {"ACT", "DIST"} and not is_trailing:
                continue
            current_params.append(c)
            continue

        # ── VÍA 3: Fallback heurístico (para datos legacy sin prefijo) ──────
        c_upper = c.upper()
        is_garbage = False
        for kw in METRIC_KEYWORDS_TO_DROP:
            if kw in c_upper:
                if any(bw in c_upper for bw in very_bad):
                    is_garbage = True; break
                if not any(ex in c_upper for ex in exceptions):
                    is_garbage = True; break
        if not is_garbage:
            current_params.append(c)

    current_params.sort()
    final_cols = current_ids + current_metrics + current_params
    return df[final_cols], current_metrics, current_params


def _ordenar_filas(df: pd.DataFrame) -> pd.DataFrame:
    if "SCORE" in df.columns:
        return df.sort_values("SCORE", ascending=False).reset_index(drop=True)
    if "SALDO_ACTUAL" in df.columns:
        return df.sort_values("SALDO_ACTUAL", ascending=False).reset_index(drop=True)
    return df


def _generar_nombre_archivo(csv, out_dir, activo, est, tf) -> str:
    fname = (f"RESUMEN_{activo}_{re.sub(r'[^A-Z0-9]', '', est.upper())}"
             f"_{re.sub(r'[^a-zA-Z0-9]', '', tf.lower())}.xlsx")
    return os.path.join(out_dir or os.path.dirname(csv), fname)


# ==============================================================================
# HELPERS OPENPYXL
# ==============================================================================

def _make_border_op(color: str, style: str = 'thin') -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _col_widths_fast_op(ws, max_col: int, max_row: int, header_row: int = 1, sample: int = 20):
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = len(str(ws.cell(row=header_row, column=col).value or ""))
        for r in range(header_row + 1, min(header_row + sample + 1, max_row + 1)):
            try:
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min((max_len + 2) * 1.15, 28)


# ==============================================================================
# ESTILOS DASHBOARD RESUMEN
# ==============================================================================

def _aplicar_estilo_avanzado(
    filepath, df, metrics_cols, params_cols, saldo_ini,
):
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import column_index_from_string
    from openpyxl.formatting.rule import CellIsRule

    wb = load_workbook(filepath)
    ws = wb.active
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column
    max_row = ws.max_row

    n_ids         = len([c for c in ID_COLS if c in df.columns])
    n_metrics     = len(metrics_cols)
    n_params      = len(params_cols)
    start_metrics = n_ids + 1
    end_metrics   = start_metrics + n_metrics - 1
    start_params  = end_metrics + 1
    end_params    = start_params + n_params - 1

    # ── Colores minimalistas ────────────────────────────────────────────────
    C_HDR      = "F7FAFC"   # fondo cabecera (gris muy claro)
    C_HDR_FONT = "718096"   # texto cabecera (gris medio)

    border_thin = _make_border_op("E2E8F0", style='thin')
    border_bot  = Border(
        bottom=Side(style='medium', color='A0AEC0'),
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
    )

    fill_hdr     = PatternFill("solid", fgColor=C_HDR)
    fill_section = PatternFill("solid", fgColor="EDF2F7")
    fill_alt     = PatternFill("solid", fgColor=COLORS["row_alt"])
    fill_w       = PatternFill("solid", fgColor="FFFFFF")

    font_section = Font(name=FONT_TITLE, size=9, bold=True, color="A0AEC0")
    font_hdr     = Font(name=FONT_TITLE, size=10, bold=True, color=C_HDR_FONT)
    font_body    = Font(name=FONT_BODY,  size=10, color=COLORS["text_dark"])
    font_id      = Font(name=FONT_BODY,  size=10, bold=True, color=COLORS["text_dark"])
    align_c      = Alignment(horizontal='center', vertical='center')
    align_l      = Alignment(horizontal='left',   vertical='center', indent=1)
    align_cw     = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Fila 1: Encabezados de sección (texto gris sutil) ───────────────────
    ws.row_dimensions[1].height = 20

    def _section(c_start, c_end, label, fill):
        c = ws.cell(row=1, column=c_start)
        c.value = label; c.fill = fill; c.font = font_section; c.alignment = align_c
        if c_end > c_start:
            ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)

    if n_ids > 0:
        _section(1, n_ids, "DATOS", fill_section)
    if n_metrics > 0:
        _section(start_metrics, end_metrics, "MÉTRICAS", fill_section)
    if n_params > 0:
        _section(start_params, end_params, "PARÁMETROS", fill_section)

    # ── Fila 2: Nombres de columna ──────────────────────────────────────────
    ws.row_dimensions[2].height = 32

    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = font_hdr; cell.alignment = align_cw
        cell.border = border_bot
        cell.fill = fill_hdr

    # ── Aplicar nombres de display en fila 2 ────────────────────────────────
    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        orig = str(cell.value or "").strip()
        if orig in METRIC_DISPLAY_NAMES:
            cell.value = METRIC_DISPLAY_NAMES[orig]

    # Leer cabeceras ya con display names
    col_hdrs = {c: str(ws.cell(2, c).value or "").strip() for c in range(1, max_col + 1)}

    # ── Filas de datos ───────────────────────────────────────────────────────
    # Columnas monetarias y porcentuales (tras renombrar)
    MONEY_HDRS = {"BEN BRUTO", "BEN NETO", "SALDO ACTUAL", "COMISIONES"}
    PCT_HDRS   = {"ROI", "MAX DD", "WIN RATE"}
    INT_HDRS   = {"LONG", "SHORT", "TRIAL"}

    for r in range(3, max_row + 1):
        rf = fill_alt if r % 2 == 0 else fill_w
        ws.row_dimensions[r].height = 22

        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            hdr  = col_hdrs[c]
            is_id_col = (c < start_metrics)

            cell.font      = font_id if is_id_col else font_body
            cell.alignment = align_l if (hdr == "ESTRATEGIA") else align_c
            cell.border    = border_thin
            cell.fill      = rf

            if hdr == "TRADES DIA":
                cell.number_format = "0.00"
            elif hdr in INT_HDRS:
                cell.number_format = "0"
            elif hdr in PCT_HDRS:
                cell.number_format = "0.00%"
                try:
                    cell.value = float(cell.value) / 100.0
                except Exception:
                    pass
            elif hdr in {"PROFIT FACTOR", "SHARPE"}:
                cell.number_format = "0.00"
            elif hdr == "EXPECTANCY":
                cell.number_format = "#,##0.00"
            elif hdr in MONEY_HDRS:
                cell.number_format = "#,##0.00"
            elif hdr == "SCORE":
                cell.number_format = "0.00"

    _col_widths_fast_op(ws, max_col, max_row, header_row=2, sample=20)

    # ── Mapa columna-display-name → letra Excel ──────────────────────────────
    col_map = {col_hdrs[c]: get_column_letter(c) for c in range(1, max_col + 1)}

    # ── Formato condicional ──────────────────────────────────────────────────
    data_start = f"3:{max_row}"

    if "BEN NETO" in col_map:
        rng = f"{col_map['BEN NETO']}3:{col_map['BEN NETO']}{max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='greaterThan', formula=['0'],
            font=Font(name=FONT_BODY, size=10, color=COLORS["accent_green"], bold=True)))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='lessThan', formula=['0'],
            font=Font(name=FONT_BODY, size=10, color=COLORS["accent_red"], bold=True)))

    ws.freeze_panes = ws.cell(row=3, column=start_metrics)

    # ── SIN GRÁFICO EN RESUMEN — La gráfica comparativa está en los trials ──

    wb.save(filepath)


# ==============================================================================
# EXPORTACIÓN RÁPIDA
# ==============================================================================

def exportar_trades_excel_rapido(
    df_trades: pd.DataFrame,
    resumen_csv_path: str,
    metrics: dict,
    params: dict,
    trial_number: int,
    trades_actual_base: str = "trades_trial",
    score: float = None,
    max_archivos: int = 5,
    skip_trades_file: bool = False,
):
    params = dict(params or {})
    fila = {
        "TRIAL":      trial_number,
        "SCORE":      score if score is not None else 0,
        "ESTRATEGIA": params.get("NOMBRE_COMBO", "UNKNOWN"),
    }
    for k, v in metrics.items():
        fila[k.upper()] = v

    def _aplanar(d, prefix=""):
        out = {}
        for k, v in d.items():
            if isinstance(v, dict):
                out.update(_aplanar(v, f"{prefix}{k.upper()}_"))
            else:
                out[f"{prefix}{k.upper()}"] = v
        return out

    fila.update(_aplanar(params))
    df_fila = pd.DataFrame([fila])

    if int(trial_number) == 0 and os.path.exists(resumen_csv_path):
        try:
            os.remove(resumen_csv_path)
        except Exception:
            pass

    mode = "w" if not os.path.exists(resumen_csv_path) else "a"
    df_fila.to_csv(resumen_csv_path, index=False, mode=mode, header=(mode == "w"))

    if not skip_trades_file:
        _gestionar_archivos_trades(df_trades, trades_actual_base, trial_number, score, max_archivos, params=params)


def _gestionar_archivos_trades(df, base_path, trial, score, max_files, params=None):
    trades_dir = os.path.dirname(base_path) or "."
    os.makedirs(trades_dir, exist_ok=True)

    s_val = score if score is not None else -999
    fpath = os.path.join(trades_dir, f"TRADES_TRIAL{trial}_SCORE{s_val:.2f}.xlsx")

    df_export = df.copy()
    for col in df_export.select_dtypes(include=["datetime64[ns, UTC]", "datetime64[ns]"]).columns:
        if hasattr(df_export[col].dt, "tz") and df_export[col].dt.tz is not None:
            df_export[col] = df_export[col].dt.tz_localize(None)

    df_export = _preparar_df_trades(df_export)

    saldo = (params or {}).get('__saldo_usado') or 0
    apal  = (params or {}).get('__apalancamiento_max') or 0

    try:
        _escribir_trades_xlsxwriter(fpath, df_export, saldo, saldo * apal, apal)
    except Exception:
        _escribir_trades_openpyxl_fallback(fpath, df_export, saldo, saldo * apal, apal)

    files = [f for f in os.listdir(trades_dir) if f.startswith("TRADES_TRIAL") and f.endswith(".xlsx")]
    if len(files) > max_files:
        scored = []
        for f in files:
            try:
                s = float(re.search(r"SCORE(-?\d+\.?\d*)", f).group(1))
                scored.append((s, f))
            except Exception:
                scored.append((-9999, f))
        scored.sort(key=lambda x: x[0], reverse=True)
        for _, f_del in scored[max_files:]:
            try:
                os.remove(os.path.join(trades_dir, f_del))
            except Exception:
                pass